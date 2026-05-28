"""
对比不同日期窗口(0, ±1, ±2, ±3, ±5, ±7)下的DFS候选池大小和计算量。
数据：Phase0 + Phase2(±7) + Phase2.5 之后剩余的各项。
"""
import openpyxl, re, math, itertools
from datetime import datetime, timedelta, date
from collections import defaultdict

MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, (int, float)) and 30000 < val < 100000:
        return (datetime(1970, 1, 1) + timedelta(days=val - 25569)).date()
    return None

def parse_amount(s):
    if s is None: return 0.0
    s = str(s).replace(',','').replace(' ','').strip()
    if not s: return 0.0
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    if s.endswith('-'): s = '-' + s[:-1]
    s = s.upper().rstrip('DR').rstrip('CR')
    try: return float(s)
    except: return 0.0

def comb_count(n, k):
    if k > n: return 0
    return math.comb(n, k)

# ==== LOAD DATA ====
BANK = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX'
ENT  = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX'
TARGET = '01469910120237'

# --- Bank ---
wb = openpyxl.load_workbook(BANK, data_only=True)
bank_raw = list(wb.active.iter_rows(values_only=True))
bank_header_raw = [str(h) if h else '' for h in bank_raw[0]]
def find_col(keyword):
    for i, h in enumerate(bank_header_raw):
        if keyword in h: return i
    raise ValueError(f"'{keyword}' not found")
acct_col = find_col('Account No')
date_col = find_col('Transaction Date')
drcr_col = find_col('Debit/Credit')
amt_col = find_col('Amount')
val_col = find_col('Value Date')
ref_col = find_col('Customer Reference')

bank_txns = []
for row in bank_raw[1:]:
    acct = str(row[acct_col]).strip().strip("'").strip() if row[acct_col] else ''
    if TARGET not in acct: continue
    raw_date = str(row[date_col]).strip() if row[date_col] else ''
    d = None
    if raw_date:
        m = re.match(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s+(\d{2}):(\d{2})', raw_date)
        if m:
            d = datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1))).date()
    if d is None:
        vd = row[val_col]
        if isinstance(vd, datetime): d = vd.date()
    if d is None: continue
    amt = parse_amount(row[amt_col])
    if amt == 0: continue
    direction = 'CR' if amt > 0 else 'DR'
    bank_txns.append({'date': d, 'amount': amt, 'direction': direction})

# --- Enterprise ---
wb = openpyxl.load_workbook(ENT, data_only=True)
ent_raw = list(wb.active.iter_rows(values_only=True))
e_date_col, e_acct_col, e_debit_col, e_credit_col = 5, 12, 15, 17

ent_txns = []
for row in ent_raw[1:]:
    if not row[e_acct_col]: continue
    acct = str(row[e_acct_col]).strip().strip("'").strip()
    if TARGET not in acct: continue
    d = parse_date(row[e_date_col])
    if not d: continue
    debit = parse_amount(row[e_debit_col])
    credit = parse_amount(row[e_credit_col])
    if debit != 0:
        ent_txns.append({'date': d, 'amount': debit, 'direction': '借'})
    elif credit != 0:
        ent_txns.append({'date': d, 'amount': credit, 'direction': '贷'})

print(f"Loaded: B{len(bank_txns)} + E{len(ent_txns)} = {len(bank_txns)+len(ent_txns)}")

# ==== PHASE 0: Account-level sum check ====
b_cr_sum = sum(t['amount'] for t in bank_txns if t['direction'] == 'CR')
e_dr_sum = sum(t['amount'] for t in ent_txns if t['direction'] == '借')
b_dr_sum = sum(t['amount'] for t in bank_txns if t['direction'] == 'DR')
e_cr_sum = sum(t['amount'] for t in ent_txns if t['direction'] == '贷')

b_cr_set = {i for i,t in enumerate(bank_txns) if t['direction'] == 'CR'}
e_dr_set = {i for i,t in enumerate(ent_txns) if t['direction'] == '借'}
b_dr_set = {i for i,t in enumerate(bank_txns) if t['direction'] == 'DR'}
e_cr_set = {i for i,t in enumerate(ent_txns) if t['direction'] == '贷'}

matched_bank = set()
matched_ent = set()

if abs(b_cr_sum - e_dr_sum) < 0.01:
    matched_bank.update(b_cr_set); matched_ent.update(e_dr_set)
    print(f"Phase0 CR: B{len(b_cr_set)}+E{len(e_dr_set)}")
if abs(b_dr_sum - e_cr_sum) < 0.01:
    matched_bank.update(b_dr_set); matched_ent.update(e_cr_set)
    print(f"Phase0 DR: B{len(b_dr_set)}+E{len(e_cr_set)}")

# ==== PHASE 2: 1:1 cents hash ±7 days ====
cents_map = defaultdict(list)
for i, t in enumerate(bank_txns):
    if i in matched_bank: continue
    key = (t['direction'], round(abs(t['amount']), 2))
    cents_map[('B', key)].append(i)
for i, t in enumerate(ent_txns):
    if i in matched_ent: continue
    e_dir = '借' if t['direction'] == '借' else '贷'
    key = (e_dir, round(abs(t['amount']), 2))
    cents_map[('E', key)].append(i)

for b_dir, e_dir in [('CR','借'), ('DR','贷')]:
    b_keys = {k[1] for (src,k), v in cents_map.items() if src=='B' and k[0]==b_dir}
    e_keys = {k[1] for (src,k), v in cents_map.items() if src=='E' and k[0]==e_dir}
    for amt in b_keys & e_keys:
        b_idx = cents_map[('B', (b_dir, amt))]
        e_idx = cents_map[('E', (e_dir, amt))]
        for bi in list(b_idx):
            if bi in matched_bank: continue
            for ei in list(e_idx):
                if ei in matched_ent: continue
                if abs((bank_txns[bi]['date'] - ent_txns[ei]['date']).days) <= 7:
                    matched_bank.add(bi); matched_ent.add(ei)

p2_count = sum(1 for i in range(len(bank_txns)) if i in matched_bank) + sum(1 for i in range(len(ent_txns)) if i in matched_ent)
print(f"Phase2 1:1: {p2_count} items matched")

# ==== PHASE 2.5: Same-date bucket ====
date_buckets = defaultdict(lambda: defaultdict(list))
for i, t in enumerate(bank_txns):
    if i in matched_bank: continue
    date_buckets[t['date']]['B_'+t['direction']].append(i)
for i, t in enumerate(ent_txns):
    if i in matched_ent: continue
    date_buckets[t['date']]['E_'+t['direction']].append(i)

p25 = 0
for d, bucket in date_buckets.items():
    for bd, ed in [('B_CR','E_借'), ('B_DR','E_贷')]:
        bi = bucket.get(bd, []); ei = bucket.get(ed, [])
        if bi and ei:
            bs = sum(bank_txns[i]['amount'] for i in bi)
            es = sum(ent_txns[i]['amount'] for i in ei)
            if abs(bs - es) < 0.01:
                for i in bi: matched_bank.add(i)
                for i in ei: matched_ent.add(i)
                p25 += len(bi) + len(ei)
print(f"Phase2.5 bucket: {p25} items matched")

# ==== REMAINING ====
remain_b = [t for i,t in enumerate(bank_txns) if i not in matched_bank]
remain_e = [t for i,t in enumerate(ent_txns) if i not in matched_ent]
print(f"\nRemaining: B{len(remain_b)} + E{len(remain_e)} = {len(remain_b)+len(remain_e)}")

# ==== WINDOW ANALYSIS ====
# For each window W, merge date groups that are within W days of each other
# Then calculate N=4,5,6 combinations per pool

all_b_dates = sorted(set(t['date'] for t in remain_b))
all_e_dates = sorted(set(t['date'] for t in remain_e))

def build_pools(remain_b, remain_e, window_days):
    """Merge items into pools where dates are within window_days of each other."""
    # Strategy: sort all unique dates, create pools where consecutive dates 
    # gap <= window_days. Then collect B and E items for each pool.
    all_dates = sorted(set(t['date'] for t in remain_b) | set(t['date'] for t in remain_e))
    
    # For each direction pair (CR/借, DR/贷), build pools
    pools = []
    for b_dir, e_dir in [('CR', '借'), ('DR', '贷')]:
        b_items = [t for t in remain_b if t['direction'] == b_dir]
        e_items = [t for t in remain_e if t['direction'] == e_dir]
        if not b_items or not e_items:
            continue
        # Get all relevant dates
        dates = sorted(set(t['date'] for t in b_items) | set(t['date'] for t in e_items))
        # Merge into pools
        if not dates:
            continue
        current_pool_dates = [dates[0]]
        pool_list = []
        for d in dates[1:]:
            if (d - current_pool_dates[-1]).days <= window_days:
                current_pool_dates.append(d)
            else:
                pool_list.append(current_pool_dates)
                current_pool_dates = [d]
        pool_list.append(current_pool_dates)
        
        for pdates in pool_list:
            dmin, dmax = min(pdates), max(pdates)
            pool_b = [t for t in b_items if dmin <= t['date'] <= dmax]
            pool_e = [t for t in e_items if dmin <= t['date'] <= dmax]
            if pool_b and pool_e:
                pools.append({
                    'dates': f"{dmin}~{dmax}" if dmin != dmax else str(dmin),
                    'nB': len(pool_b),
                    'nE': len(pool_e),
                    'dir': f"{b_dir}/{e_dir}"
                })
    return pools

print("\n" + "=" * 100)
print(f"{'Window':>7} {'Pools':>6} {'Max B':>6} {'Max E':>6} {'Max items':>10} {'N=4 combos':>14} {'N=5 combos':>16} {'N=6 combos':>18}")
print("=" * 100)

for window in [0, 1, 2, 3, 5, 7]:
    pools = build_pools(remain_b, remain_e, window)
    
    max_b = max(p['nB'] for p in pools) if pools else 0
    max_e = max(p['nE'] for p in pools) if pools else 0
    max_items = max(p['nB'] + p['nE'] for p in pools) if pools else 0
    
    # Find the worst pool
    worst = max(pools, key=lambda p: p['nB'] + p['nE']) if pools else None
    
    n4 = n5 = n6 = 0
    for p in pools:
        nb, ne = p['nB'], p['nE']
        # 1:B = N:E → nb * C(ne, N)
        # N:B = 1:E → ne * C(nb, N)  
        n4 += nb * comb_count(ne, 4) + ne * comb_count(nb, 4)
        n5 += nb * comb_count(ne, 5) + ne * comb_count(nb, 5)
        n6 += nb * comb_count(ne, 6) + ne * comb_count(nb, 6)
        # Also count 2:2, 3:2, 2:3 etc which are significant
        n4 += comb_count(nb, 2) * comb_count(ne, 2)  # 2:2
        n5 += comb_count(nb, 2) * comb_count(ne, 3) + comb_count(nb, 3) * comb_count(ne, 2)  # 2:3 + 3:2
        n6 += (comb_count(nb, 2) * comb_count(ne, 4) + comb_count(nb, 3) * comb_count(ne, 3) + 
               comb_count(nb, 4) * comb_count(ne, 2))  # 2:4 + 3:3 + 4:2
    
    label = f"±{window}" if window > 0 else "strict"
    print(f"{label:>7} {len(pools):>6} {max_b:>6} {max_e:>6} {max_items:>10} {n4:>14,} {n5:>16,} {n6:>18,}")
    
    # Show worst pool details
    if worst:
        nb, ne = worst['nB'], worst['nE']
        n4_w = nb * comb_count(ne,4) + ne * comb_count(nb,4) + comb_count(nb,2)*comb_count(ne,2)
        n5_w = nb * comb_count(ne,5) + ne * comb_count(nb,5) + comb_count(nb,2)*comb_count(ne,3) + comb_count(nb,3)*comb_count(ne,2)
        n6_w = nb * comb_count(ne,6) + ne * comb_count(nb,6) + comb_count(nb,2)*comb_count(ne,4) + comb_count(nb,3)*comb_count(ne,3) + comb_count(nb,4)*comb_count(ne,2)
        print(f"       worst pool: {worst['dates']} B{nb}+E{ne} {worst['dir']}  N4={n4_w:,} N5={n5_w:,} N6={n6_w:,}")

print("=" * 100)

# ==== QUICK SUMMARY ====
print("\n=== N=6 total combos comparison ===")
for window in [0, 1, 2, 3, 5, 7]:
    pools = build_pools(remain_b, remain_e, window)
    n6 = 0
    for p in pools:
        nb, ne = p['nB'], p['nE']
        n6 += nb * comb_count(ne, 6) + ne * comb_count(nb, 6)
        n6 += comb_count(nb,2)*comb_count(ne,4) + comb_count(nb,3)*comb_count(ne,3) + comb_count(nb,4)*comb_count(ne,2)
    label = f"±{window}" if window > 0 else "strict"
    print(f"  {label:>7}: {n6:>20,}")
