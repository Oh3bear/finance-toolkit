#!/usr/bin/env python3
"""分析剩余283条的匹配障碍：放宽约束能捡回多少"""
import openpyxl, re, io, os
from datetime import datetime, timedelta
from collections import defaultdict

OUTPUT = r'C:\Users\CSCI\Downloads\kimiOKC\财务工具集\recon_relax.txt'
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def parse_date(val):
    if isinstance(val, datetime): return val
    if isinstance(val, (int, float)) and 30000 < val < 100000:
        return datetime(1970, 1, 1) + timedelta(days=val - 25569)
    return None

def to_date_key(d): return '%d-%02d-%02d' % (d.year, d.month, d.day)

def cents_eq(a, b): return round(a * 100) == round(b * 100)

def within_window(a, b, days=7):
    return abs((a - b).total_seconds()) <= days * 86400

def parse_amount(s):
    if s is None: return 0.0
    s = str(s).replace(',','').replace(' ','').strip()
    if not s: return 0.0
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    if s.endswith('-'): s = '-' + s[:-1]
    s = s.upper().rstrip('DR').rstrip('CR')
    try: return float(s)
    except: return 0.0

BANK = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX'
ENT  = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX'

# ---- read bank ----
wb = openpyxl.load_workbook(BANK, data_only=True)
bank_raw = list(wb.active.iter_rows(values_only=True))
bank_header = [str(h).replace('\n',' ').strip() if h else '' for h in bank_raw[0]]
acct_col = next(i for i,h in enumerate(bank_header) if 'Account No' in h)
date_col = next(i for i,h in enumerate(bank_header) if 'Transaction Date' in h)
drcr_col = next(i for i,h in enumerate(bank_header) if 'Debit/Credit' in h)
amt_col = next(i for i,h in enumerate(bank_header) if 'Amount' in h)
val_col = next(i for i,h in enumerate(bank_header) if 'Value Date' in h)

bank_txns = []
for row in bank_raw[1:]:
    acct = str(row[acct_col]).strip().strip("'").strip() if row[acct_col] else ''
    if not acct: continue
    raw_date = str(row[date_col]).strip() if row[date_col] else ''
    if not raw_date:
        vd = row[val_col]
        if isinstance(vd, datetime): d = vd
        else: continue
    else:
        m = re.match(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s+(\d{2}):(\d{2})', raw_date)
        if m:
            d = datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1)), int(m.group(4)), int(m.group(5)))
        else:
            vd = row[val_col]
            if isinstance(vd, datetime): d = vd
            else: continue
    amt = parse_amount(row[amt_col])
    drcr = str(row[drcr_col]).strip().upper() if row[drcr_col] else ''
    if amt == 0: continue
    direction = 'CR' if drcr == 'CR' else 'DR'
    amount = abs(amt) if drcr == 'CR' else -abs(amt)
    bank_txns.append({'account': acct, 'date': d, 'amount': amount, 'direction': direction, 'desc': str(row[12])[:40]})

# ---- read enterprise ----
wb = openpyxl.load_workbook(ENT, data_only=True)
ent_raw = list(wb.active.iter_rows(values_only=True))
ent_txns = []
e_acct_col, e_date_col, e_debit_col, e_credit_col = 12, 5, 15, 17
for row in ent_raw[1:]:
    if not row[e_acct_col]: continue
    acct = str(row[e_acct_col]).strip().strip("'")
    if not acct: continue
    d = parse_date(row[e_date_col])
    if not d: continue
    debit = parse_amount(row[e_debit_col]) if len(row) > e_debit_col else 0
    credit = parse_amount(row[e_credit_col]) if len(row) > e_credit_col else 0
    if debit != 0:
        ent_txns.append({'account': acct, 'date': d, 'amount': debit, 'direction': 'borrow', 'desc': str(row[8])[:40]})
    elif credit != 0:
        ent_txns.append({'account': acct, 'date': d, 'amount': credit, 'direction': 'lend', 'desc': str(row[8])[:40]})

# ---- Phase 0 + Phase 2 + Phase 2.5 (same as before) ----
bank_by_acct = defaultdict(list)
for t in bank_txns: bank_by_acct[t['account']].append(t)
ent_by_acct = defaultdict(list)
for t in ent_txns: ent_by_acct[t['account']].append(t)

all_remaining_b = []
all_remaining_e = []

for acct in sorted(set(bank_by_acct.keys()) & set(ent_by_acct.keys())):
    b_list = bank_by_acct[acct]
    e_list = ent_by_acct[acct]

    b_inc = [t for t in b_list if t['direction'] == 'CR']
    b_exp = [t for t in b_list if t['direction'] == 'DR']
    e_deb = [t for t in e_list if t['direction'] == 'borrow']
    e_cre = [t for t in e_list if t['direction'] == 'lend']

    b_inc_sum = sum(t['amount'] for t in b_inc)
    b_exp_sum = sum(t['amount'] for t in b_exp)
    e_deb_sum = sum(t['amount'] for t in e_deb)
    e_cre_sum = sum(t['amount'] for t in e_cre)

    used_b = set(); used_e = set()

    if cents_eq(b_inc_sum, e_deb_sum) and b_inc and e_deb:
        for i in range(len(b_list)):
            if b_list[i] in b_inc: used_b.add(i)
        for i in range(len(e_list)):
            if e_list[i] in e_deb: used_e.add(i)
    if cents_eq(b_exp_sum, e_cre_sum) and b_exp and e_cre:
        for i in range(len(b_list)):
            if b_list[i] in b_exp: used_b.add(i)
        for i in range(len(e_list)):
            if e_list[i] in e_cre: used_e.add(i)

    # Phase 2: 1:1 by cents + date window
    ent_by_cents = defaultdict(list)
    for i, e in enumerate(e_list):
        if i in used_e: continue
        ent_by_cents[round(e['amount'] * 100)].append(i)
    for bi, b in enumerate(b_list):
        if bi in used_b: continue
        c = round(b['amount'] * 100)
        bucket = ent_by_cents.get(c, [])
        if not bucket: continue
        best = None
        for ei in bucket:
            if ei in used_e: continue
            if within_window(b['date'], e_list[ei]['date']):
                best = ei; break
        if best is None:
            for ei in bucket:
                if ei not in used_e: best = ei; break
        if best is not None:
            used_b.add(bi); used_e.add(best)

    # Phase 2.5: date bucket
    remaining_b = [(i, b_list[i]) for i in range(len(b_list)) if i not in used_b]
    remaining_e = [(i, e_list[i]) for i in range(len(e_list)) if i not in used_e]
    b_by_date = defaultdict(list)
    for idx, t in remaining_b: b_by_date[to_date_key(t['date'])].append((idx, t))
    e_by_date = defaultdict(list)
    for idx, t in remaining_e: e_by_date[to_date_key(t['date'])].append((idx, t))
    for dk, b_items in b_by_date.items():
        e_items = e_by_date.get(dk, [])
        b_inc_d = [(i,t) for i,t in b_items if t['amount'] > 0]
        e_deb_d = [(i,t) for i,t in e_items if t['amount'] > 0 and t['direction'] == 'borrow']
        if b_inc_d and e_deb_d:
            bs = sum(t['amount'] for _,t in b_inc_d)
            es = sum(t['amount'] for _,t in e_deb_d)
            if cents_eq(bs, es):
                for i,_ in b_inc_d: used_b.add(i)
                for i,_ in e_deb_d: used_e.add(i)
        b_exp_d = [(i,t) for i,t in b_items if t['amount'] < 0]
        e_cre_d = [(i,t) for i,t in e_items if t['amount'] < 0 and t['direction'] == 'lend']
        if b_exp_d and e_cre_d:
            bs = sum(t['amount'] for _,t in b_exp_d)
            es = sum(t['amount'] for _,t in e_cre_d)
            if cents_eq(bs, es):
                for i,_ in b_exp_d: used_b.add(i)
                for i,_ in e_cre_d: used_e.add(i)

    rem_b = [(i, b_list[i]) for i in range(len(b_list)) if i not in used_b]
    rem_e = [(i, e_list[i]) for i in range(len(e_list)) if i not in used_e]
    all_remaining_b.extend(rem_b)
    all_remaining_e.extend(rem_e)

# ---- Now analyze the remaining items ----
out = io.StringIO()
out.write('===== 剩余 B%d+E%d 障碍分析 =====\n\n' % (len(all_remaining_b), len(all_remaining_e)))

# 1. Same-cents check: items where the exact same cent value exists but outside date window
rem_b_cents = defaultdict(list)
for i, t in all_remaining_b:
    rem_b_cents[round(t['amount'] * 100)].append((i, t))
rem_e_cents = defaultdict(list)
for i, t in all_remaining_e:
    rem_e_cents[round(t['amount'] * 100)].append((i, t))

# 2. Same-cents B↔E pairs that exist but exceed date window
date_mismatch = []
for cents_key in set(rem_b_cents.keys()) & set(rem_e_cents.keys()):
    for bi, bt in rem_b_cents[cents_key]:
        for ei, et in rem_e_cents[cents_key]:
            diff_days = abs((bt['date'] - et['date']).days)
            if diff_days > 7:
                date_mismatch.append((diff_days, bi, bt, ei, et))

date_mismatch.sort()
out.write('--- 1. 同金额但超出 ±7天窗口 ---\n')
out.write('共 %d 对候选\n' % len(date_mismatch))
by_range = defaultdict(int)
for dd, _, _, _, _ in date_mismatch:
    if dd <= 14: by_range['8-14天'] += 1
    elif dd <= 30: by_range['15-30天'] += 1
    else: by_range['>30天'] += 1
for k in ['8-14天', '15-30天', '>30天']:
    out.write('  %s: %d 对\n' % (k, by_range.get(k, 0)))

# Show some examples
out.write('\n  示例:\n')
for dd, bi, bt, ei, et in date_mismatch[:5]:
    out.write('  日期差%d天: B=%s %.2f | E=%s %.2f\n' % (
        dd, to_date_key(bt['date']), bt['amount'],
        to_date_key(et['date']), et['amount']))

# 3. Cents close but not exact (difference of 1-2 cents)
close_cents = []
all_e_amt = [(round(t['amount'] * 100), i, t) for i, t in all_remaining_e]
all_e_amt.sort()
all_b_amt = [(round(t['amount'] * 100), i, t) for i, t in all_remaining_b]

for bc, bi, bt in all_b_amt:
    # Binary search nearby
    for ec, ei, et in all_e_amt:
        diff = abs(bc - ec)
        if diff > 2: continue
        if not within_window(bt['date'], et['date'], 7): continue
        if diff > 0:
            close_cents.append((diff, bi, bt, ei, et))
close_cents.sort()

out.write('\n--- 2. 金额差 ±2分 且日期在窗口内 ---\n')
out.write('共 %d 对候选\n' % len(close_cents))
by_diff = defaultdict(int)
for d, _, _, _, _ in close_cents:
    by_diff['%d分' % d] += 1
for k in sorted(by_diff):
    out.write('  %s: %d 对\n' % (k, by_diff[k]))

out.write('\n  示例:\n')
for diff, bi, bt, ei, et in close_cents[:5]:
    out.write('  差%d分: B=%s %.2f | E=%s %.2f (描述: B=%s | E=%s)\n' % (
        diff, to_date_key(bt['date']), bt['amount'],
        to_date_key(et['date']), et['amount'],
        bt.get('desc','')[:30], et.get('desc','')[:30]))

# 4. Remaining items distribution by direction and amount
out.write('\n--- 3. 剩余条目按方向分布 ---\n')
b_cr = [t for _, t in all_remaining_b if t['amount'] > 0]
b_dr = [t for _, t in all_remaining_b if t['amount'] < 0]
e_deb = [t for _, t in all_remaining_e if t['amount'] > 0]
e_cre = [t for _, t in all_remaining_e if t['amount'] < 0]
out.write('B-CR(收入): %d 笔, 合计 %.2f\n' % (len(b_cr), sum(t['amount'] for t in b_cr)))
out.write('B-DR(支出): %d 笔, 合计 %.2f\n' % (len(b_dr), sum(t['amount'] for t in b_dr)))
out.write('E-借(收入): %d 笔, 合计 %.2f\n' % (len(e_deb), sum(t['amount'] for t in e_deb)))
out.write('E-贷(支出): %d 笔, 合计 %.2f\n' % (len(e_cre), sum(t['amount'] for t in e_cre)))

# 5. Amount magnitude distribution
out.write('\n--- 4. 剩余条目按金额量级分布 ---\n')
def magnitude(amt):
    a = abs(amt)
    if a >= 10000000: return '千万'
    if a >= 1000000: return '百万'
    if a >= 100000: return '十万'
    if a >= 10000: return '万'
    if a >= 1000: return '千'
    return '百以下'

for label, items in [('B', all_remaining_b), ('E', all_remaining_e)]:
    dist = defaultdict(int)
    for _, t in items:
        dist[magnitude(t['amount'])] += 1
    out.write('  %s: %s\n' % (label, dict(sorted(dist.items()))))

# 6. Unique cent count analysis
out.write('\n--- 5. 唯一分位数分析 ---\n')
b_cents = set(round(t['amount'] * 100) for _, t in all_remaining_b)
e_cents = set(round(t['amount'] * 100) for _, t in all_remaining_e)
overlap = b_cents & e_cents
out.write('B 唯一分位数: %d / %d (%.0f%%)\n' % (len(b_cents), len(all_remaining_b), 
    len(b_cents)/len(all_remaining_b)*100 if all_remaining_b else 0))
out.write('E 唯一分位数: %d / %d (%.0f%%)\n' % (len(e_cents), len(all_remaining_e),
    len(e_cents)/len(all_remaining_e)*100 if all_remaining_e else 0))
out.write('重叠分位数: %d\n' % len(overlap))

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(out.getvalue())
print(out.getvalue())
