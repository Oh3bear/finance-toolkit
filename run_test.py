#!/usr/bin/env python3
"""银企对账模拟 - 输出到工作目录"""
import openpyxl, re, io, os
from datetime import datetime, timedelta
from collections import defaultdict

OUTPUT = r'C:\Users\CSCI\Downloads\kimiOKC\财务工具集\recon_result.txt'

os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

out = io.StringIO()

MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def parse_date(val):
    if isinstance(val, datetime): return val
    if isinstance(val, (int, float)) and 30000 < val < 100000:
        return datetime(1970, 1, 1) + timedelta(days=val - 25569)
    return None

def to_date_key(d):
    return '%d-%02d-%02d' % (d.year, d.month, d.day)

def cents_eq(a, b): return round(a * 100) == round(b * 100)

def within_window(a, b, days=7):
    return abs((a - b).total_seconds()) <= days * 86400

def parse_amount(s):
    if s is None: return 0.0
    s = str(s).replace(',','').replace(' ','').strip()
    if not s: return 0.0
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    if s.endswith('-'): s = '-' + s[:-1]
    s = s.upper().rstrip('DR').rstrip('CR')
    try: return float(s)
    except: return 0.0

BANK = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX'
ENT  = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX'

# read bank
wb = openpyxl.load_workbook(BANK, data_only=True)
bank_raw = list(wb.active.iter_rows(values_only=True))
bank_header = [str(h).replace('\n',' ').strip() if h else '' for h in bank_raw[0]]
acct_col = next(i for i,h in enumerate(bank_header) if 'Account No' in h)
date_col = next(i for i,h in enumerate(bank_header) if 'Transaction Date' in h)
drcr_col = next(i for i,h in enumerate(bank_header) if 'Debit/Credit' in h)
amt_col = next(i for i,h in enumerate(bank_header) if h.strip() == 'Amount')
val_col = next(i for i,h in enumerate(bank_header) if 'Value Date' in h)

bank_txns = []
for row in bank_raw[1:]:
    acct = str(row[acct_col]).strip().strip("'").strip() if row[acct_col] else ''
    if not acct: continue
    raw_date = str(row[date_col]).strip() if row[date_col] else ''
    if not raw_date:
        vd = row[val_col]
        if isinstance(vd, datetime): d = vd
        else: continue
    else:
        m = re.match(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s+(\d{2}):(\d{2})', raw_date)
        if m:
            d = datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1)), int(m.group(4)), int(m.group(5)))
        else:
            vd = row[val_col]
            if isinstance(vd, datetime): d = vd
            else: continue
    amt = parse_amount(row[amt_col])
    drcr = str(row[drcr_col]).strip().upper() if row[drcr_col] else ''
    if amt == 0: continue
    direction = 'CR' if drcr == 'CR' else 'DR'
    amount = abs(amt) if drcr == 'CR' else -abs(amt)
    bank_txns.append({'account': acct, 'date': d, 'amount': amount, 'direction': direction})

out.write('Bank: %d txns\n' % len(bank_txns))
out.write('Bank dates: %s ~ %s\n' % (min(t['date'] for t in bank_txns), max(t['date'] for t in bank_txns)))
out.write('Bank samples: %s\n' % [to_date_key(t['date']) for t in bank_txns[:5]])

# read enterprise
wb = openpyxl.load_workbook(ENT, data_only=True)
ent_raw = list(wb.active.iter_rows(values_only=True))
ent_header = [str(h).replace('\n',' ').strip() if h else '' for h in ent_raw[0]]

e_acct_col = 12
e_date_col = 5
e_debit_col = 15
e_credit_col = 17

ent_txns = []
for row in ent_raw[1:]:
    if not row[e_acct_col]: continue
    acct = str(row[e_acct_col]).strip().strip("'")
    if not acct: continue
    d = parse_date(row[e_date_col])
    if not d: continue
    debit = parse_amount(row[e_debit_col]) if len(row) > e_debit_col else 0
    credit = parse_amount(row[e_credit_col]) if len(row) > e_credit_col else 0
    if debit != 0:
        ent_txns.append({'account': acct, 'date': d, 'amount': debit, 'direction': 'borrow'})
    elif credit != 0:
        ent_txns.append({'account': acct, 'date': d, 'amount': -credit, 'direction': 'lend'})

out.write('Ent: %d txns\n' % len(ent_txns))
out.write('Ent dates: %s ~ %s\n' % (min(t['date'] for t in ent_txns), max(t['date'] for t in ent_txns)))
out.write('Ent samples: %s\n' % [to_date_key(t['date']) for t in ent_txns[:10]])

# account grouping
bank_by_acct = defaultdict(list)
for t in bank_txns: bank_by_acct[t['account']].append(t)
ent_by_acct = defaultdict(list)
for t in ent_txns: ent_by_acct[t['account']].append(t)

bank_accts = sorted(bank_by_acct.keys())
ent_accts = sorted(ent_by_acct.keys())
overlap = sorted(set(bank_accts) & set(ent_accts))

out.write('\nAccounts:\n')
out.write('  Bank: %s\n' % bank_accts)
out.write('  Ent:  %s\n' % ent_accts)
out.write('  Overlap: %s\n' % overlap)

# match
total_bank = 0; total_ent = 0; total_1to1 = 0; total_quick = 0; total_bucket = 0
total_rem_b = 0; total_rem_e = 0

for acct in overlap:
    b_list = sorted(bank_by_acct[acct], key=lambda t: t['date'])
    e_list = sorted(ent_by_acct[acct], key=lambda t: t['date'])
    total_bank += len(b_list); total_ent += len(e_list)
    
    b_inc = [t for t in b_list if t['direction'] == 'CR']
    b_exp = [t for t in b_list if t['direction'] == 'DR']
    e_deb = [t for t in e_list if t['direction'] == 'borrow']
    e_cre = [t for t in e_list if t['direction'] == 'lend']
    
    b_inc_sum = sum(t['amount'] for t in b_inc)
    b_exp_sum = sum(t['amount'] for t in b_exp)
    e_deb_sum = sum(t['amount'] for t in e_deb)
    e_cre_sum = sum(t['amount'] for t in e_cre)
    
    fast_inc = cents_eq(b_inc_sum, e_deb_sum) and b_inc and e_deb
    fast_exp = cents_eq(b_exp_sum, e_cre_sum) and b_exp and e_cre
    
    used_b = set(); used_e = set()
    
    if fast_inc:
        for t in b_inc: used_b.add(b_list.index(t))
        for t in e_deb: used_e.add(e_list.index(t))
        total_quick += len(b_inc) + len(e_deb)
    if fast_exp:
        for t in b_exp: used_b.add(b_list.index(t))
        for t in e_cre: used_e.add(e_list.index(t))
        total_quick += len(b_exp) + len(e_cre)
    
    # Phase 2: 1:1
    ent_by_cents = defaultdict(list)
    for i, e in enumerate(e_list):
        if i in used_e: continue
        ent_by_cents[round(e['amount'] * 100)].append(i)
    
    matched = 0
    for bi, b in enumerate(b_list):
        if bi in used_b: continue
        c = round(b['amount'] * 100)
        bucket = ent_by_cents.get(c, [])
        if not bucket: continue
        best = None
        for ei in bucket:
            if ei in used_e: continue
            if within_window(b['date'], e_list[ei]['date']):
                best = ei; break
        if best is None:
            for ei in bucket:
                if ei not in used_e:
                    best = ei; break
        if best is not None:
            used_b.add(bi); used_e.add(best); matched += 1
    total_1to1 += matched
    
    # Phase 2.5
    remaining_b = [(i, b_list[i]) for i in range(len(b_list)) if i not in used_b]
    remaining_e = [(i, e_list[i]) for i in range(len(e_list)) if i not in used_e]
    
    b_by_date = defaultdict(list)
    for idx, t in remaining_b: b_by_date[to_date_key(t['date'])].append((idx, t))
    e_by_date = defaultdict(list)
    for idx, t in remaining_e: e_by_date[to_date_key(t['date'])].append((idx, t))
    
    bucket_matched = 0
    for dk, b_items in b_by_date.items():
        e_items = e_by_date.get(dk, [])
        b_inc_d = [(i,t) for i,t in b_items if t['amount'] > 0]
        e_deb_d = [(i,t) for i,t in e_items if t['amount'] > 0 and t['direction'] == 'borrow']
        if b_inc_d and e_deb_d:
            bs = sum(t['amount'] for _,t in b_inc_d)
            es = sum(t['amount'] for _,t in e_deb_d)
            if cents_eq(bs, es):
                for i,_ in b_inc_d: used_b.add(i)
                for i,_ in e_deb_d: used_e.add(i)
                bucket_matched += len(b_inc_d) + len(e_deb_d)
        b_exp_d = [(i,t) for i,t in b_items if t['amount'] < 0]
        e_cre_d = [(i,t) for i,t in e_items if t['amount'] < 0 and t['direction'] == 'lend']
        if b_exp_d and e_cre_d:
            bs = sum(t['amount'] for _,t in b_exp_d)
            es = sum(t['amount'] for _,t in e_cre_d)
            if cents_eq(bs, es):
                for i,_ in b_exp_d: used_b.add(i)
                for i,_ in e_cre_d: used_e.add(i)
                bucket_matched += len(b_exp_d) + len(e_cre_d)
    total_bucket += bucket_matched
    
    rem_b = len([i for i in range(len(b_list)) if i not in used_b])
    rem_e = len([i for i in range(len(e_list)) if i not in used_e])
    total_rem_b += rem_b; total_rem_e += rem_e
    
    flags = []
    if fast_inc: flags.append('FI')
    if fast_exp: flags.append('FE')
    status = 'OK' if rem_b == 0 and rem_e == 0 else 'LEFT%d+%d' % (rem_b, rem_e)
    out.write('\n%s: B%d E%d Fast:%s 1:1:%d Bkt:%d %s\n' % (
        acct, len(b_list), len(e_list), ','.join(flags) if flags else '-',
        matched, bucket_matched, status))
    out.write('  sums: inc=%.2f/%.2f(diff=%dc) exp=%.2f/%.2f(diff=%dc)\n' % (
        b_inc_sum, e_deb_sum, round((b_inc_sum-e_deb_sum)*100),
        b_exp_sum, e_cre_sum, round((b_exp_sum-e_cre_sum)*100)))

out.write('\n===== SUMMARY =====\n')
out.write('Total: Bank %d | Ent %d\n' % (total_bank, total_ent))
out.write('Phase0 Fast: %d items (%d pairs)\n' % (total_quick, total_quick//2))
out.write('Phase2 1:1: %d pairs\n' % total_1to1)
out.write('Phase2.5 Bucket: %d items\n' % total_bucket)
out.write('Remaining: Bank %d | Ent %d\n' % (total_rem_b, total_rem_e))
matched_all = total_bank + total_ent - total_rem_b - total_rem_e
rate = matched_all / (total_bank + total_ent) * 100 if (total_bank + total_ent) > 0 else 0
out.write('Match rate: %.1f%% (%d/%d)\n' % (rate, matched_all, total_bank+total_ent))

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(out.getvalue())
print('DONE')
