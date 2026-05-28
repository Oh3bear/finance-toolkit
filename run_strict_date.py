"""
严格按日期分组，分析剩余283条的N≤4到无上限匹配计算量。
数据源：Phase0 + Phase2 + Phase2.5 已匹配后剩余的 B52 + E231。
"""
import openpyxl, re, math, itertools
from datetime import datetime, timedelta
from collections import defaultdict

MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, (int, float)) and 30000 < val < 100000:
        return (datetime(1970, 1, 1) + timedelta(days=val - 25569)).date()
    return None

def parse_amount(s):
    if s is None: return 0.0
    s = str(s).replace(',','').replace(' ','').strip()
    if not s: return 0.0
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    if s.endswith('-'): s = '-' + s[:-1]
    s = s.upper().rstrip('DR').rstrip('CR')
    try: return float(s)
    except: return 0.0

def comb_count(n, k):
    """C(n,k)"""
    if k > n: return 0
    return math.comb(n, k)

# ==== LOAD DATA ====
BANK = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX'
ENT  = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX'

TARGET = '01469910120237'

# --- Bank ---
wb = openpyxl.load_workbook(BANK, data_only=True)
bank_raw = list(wb.active.iter_rows(values_only=True))
bank_header_raw = [str(h) if h else '' for h in bank_raw[0]]
def find_col(keyword):
    for i, h in enumerate(bank_header_raw):
        if keyword in h:
            return i
    raise ValueError(f"Cannot find '{keyword}' in bank header")

acct_col = find_col('Account No')
date_col = find_col('Transaction Date')
drcr_col = find_col('Debit/Credit')
amt_col = find_col('Amount')
val_col = find_col('Value Date')
ref_col = find_col('Customer Reference')

bank_txns = []
for row in bank_raw[1:]:
    acct = str(row[acct_col]).strip().strip("'").strip() if row[acct_col] else ''
    if TARGET not in acct: continue
    raw_date = str(row[date_col]).strip() if row[date_col] else ''
    d = None
    if raw_date:
        m = re.match(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s+(\d{2}):(\d{2})', raw_date)
        if m:
            d = datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1))).date()
    if d is None:
        vd = row[val_col]
        if isinstance(vd, datetime): d = vd.date()
    if d is None: continue
    amt = parse_amount(row[amt_col])
    if amt == 0: continue
    ref = str(row[ref_col]).strip().strip("'").strip() if row[ref_col] else ''
    if amt > 0:
        direction = 'CR'  # 银行收入
    else:
        direction = 'DR'  # 银行支出
    bank_txns.append({'date': d, 'amount': amt, 'direction': direction, 'ref': ref[:60]})

# --- Enterprise ---
wb = openpyxl.load_workbook(ENT, data_only=True)
ent_raw = list(wb.active.iter_rows(values_only=True))
# Column indices from prior exploration: 5=记账日期, 12=资金账户, 15=借方, 17=贷方, 11=摘要
e_date_col = 5
e_acct_col = 12
e_debit_col = 15
e_credit_col = 17
e_memo_col = 11

ent_txns = []
for row in ent_raw[1:]:
    if not row[e_acct_col]: continue
    acct = str(row[e_acct_col]).strip().strip("'").strip()
    if TARGET not in acct: continue
    d = parse_date(row[e_date_col])
    if not d: continue
    debit = parse_amount(row[e_debit_col]) if len(row) > e_debit_col else 0
    credit = parse_amount(row[e_credit_col]) if len(row) > e_credit_col else 0
    memo = str(row[e_memo_col]).strip().strip("'").strip()[:60] if row[e_memo_col] else ''
    if debit != 0:
        ent_txns.append({'date': d, 'amount': debit, 'direction': '借'})
    elif credit != 0:
        ent_txns.append({'date': d, 'amount': credit, 'direction': '贷'})

# ==== RUN PHASES 0, 2, 2.5 ====
# Phase 0: Account-level sum matching
# -- Income (CR) --
b_cr_sum = sum(t['amount'] for t in bank_txns if t['direction'] == 'CR')
e_dr_sum = sum(t['amount'] for t in ent_txns if t['direction'] == '借')
b_cr_set = {i for i,t in enumerate(bank_txns) if t['direction'] == 'CR'}
e_dr_set = {i for i,t in enumerate(ent_txns) if t['direction'] == '借'}
cr_balanced = abs(b_cr_sum - e_dr_sum) < 0.01

# -- Expense (DR) --
b_dr_sum = sum(t['amount'] for t in bank_txns if t['direction'] == 'DR')
e_cr_sum = sum(t['amount'] for t in ent_txns if t['direction'] == '贷')
b_dr_set = {i for i,t in enumerate(bank_txns) if t['direction'] == 'DR'}
e_cr_set = {i for i,t in enumerate(ent_txns) if t['direction'] == '贷'}
dr_balanced = abs(b_dr_sum - e_cr_sum) < 0.01

print(f"Phase0 CR: bank={b_cr_sum:.2f} ent={e_dr_sum:.2f} diff={abs(b_cr_sum-e_dr_sum):.4f} balanced={cr_balanced}")
print(f"Phase0 DR: bank={b_dr_sum:.2f} ent={e_cr_sum:.2f} diff={abs(b_dr_sum-e_cr_sum):.4f} balanced={dr_balanced}")

matched_bank = set()
matched_ent = set()

p0_matched = 0
if cr_balanced:
    matched_bank.update(b_cr_set)
    matched_ent.update(e_dr_set)
    p0_matched += len(b_cr_set) + len(e_dr_set)
    print(f"Phase0 CR matched: B{len(b_cr_set)} + E{len(e_dr_set)}")

if dr_balanced:
    matched_bank.update(b_dr_set)
    matched_ent.update(e_cr_set)
    p0_matched += len(b_dr_set) + len(e_cr_set)
    print(f"Phase0 DR matched: B{len(b_dr_set)} + E{len(e_cr_set)}")

# Phase 2: 1:1 cents hash with ±7 day window
cents_map = defaultdict(list)  # keyed by (direction, absolute_cents)
for i, t in enumerate(bank_txns):
    if i in matched_bank: continue
    key = (t['direction'], round(abs(t['amount']), 2))
    cents_map[('B', key)].append(i)

for i, t in enumerate(ent_txns):
    if i in matched_ent: continue
    e_dir = '借' if t['direction'] == '借' else '贷'
    key = (e_dir, round(abs(t['amount']), 2))
    cents_map[('E', key)].append(i)

p2_matched = 0
# CR ↔ 借, DR ↔ 贷
dir_map = {('CR','借'), ('DR','贷')}
for b_dir, e_dir in dir_map:
    # Collect all unique amount keys
    b_keys = set()
    e_keys = set()
    for (src, key), indices in cents_map.items():
        if src == 'B' and key[0] == b_dir:
            b_keys.add(key[1])
        elif src == 'E' and key[0] == e_dir:
            e_keys.add(key[1])
    
    common = b_keys & e_keys
    for amt in common:
        b_indices = cents_map[('B', (b_dir, amt))]
        e_indices = cents_map[('E', (e_dir, amt))]
        # Date window filter
        for bi in list(b_indices):
            bd = bank_txns[bi]['date']
            for ei in list(e_indices):
                ed = ent_txns[ei]['date']
                if abs((bd - ed).days) <= 7:
                    if bi not in matched_bank and ei not in matched_ent:
                        matched_bank.add(bi)
                        matched_ent.add(ei)
                        p2_matched += 2
print(f"Phase2 1:1 matched: {p2_matched} items")

# Phase 2.5: Same-date bucket aggregation
date_buckets = defaultdict(lambda: defaultdict(list))  # date -> direction -> [indices]
for i, t in enumerate(bank_txns):
    if i in matched_bank: continue
    date_buckets[t['date']]['B_'+t['direction']].append(i)
for i, t in enumerate(ent_txns):
    if i in matched_ent: continue
    date_buckets[t['date']]['E_'+t['direction']].append(i)

p25_matched = 0
for d, bucket in date_buckets.items():
    # CR vs 借
    b_cr_idx = bucket.get('B_CR', [])
    e_dr_idx = bucket.get('E_借', [])
    if b_cr_idx and e_dr_idx:
        b_sum = sum(bank_txns[i]['amount'] for i in b_cr_idx)
        e_sum = sum(ent_txns[i]['amount'] for i in e_dr_idx)
        if abs(b_sum - e_sum) < 0.01:
            for idx in b_cr_idx: matched_bank.add(idx)
            for idx in e_dr_idx: matched_ent.add(idx)
            p25_matched += len(b_cr_idx) + len(e_dr_idx)
    # DR vs 贷
    b_dr_idx = bucket.get('B_DR', [])
    e_cr_idx = bucket.get('E_贷', [])
    if b_dr_idx and e_cr_idx:
        b_sum = sum(bank_txns[i]['amount'] for i in b_dr_idx)
        e_sum = sum(ent_txns[i]['amount'] for i in e_cr_idx)
        if abs(b_sum - e_sum) < 0.01:
            for idx in b_dr_idx: matched_bank.add(idx)
            for idx in e_cr_idx: matched_ent.add(idx)
            p25_matched += len(b_dr_idx) + len(e_cr_idx)
print(f"Phase2.5 bucket matched: {p25_matched} items")

# ==== REMAINING ====
remain_bank = []
for i, t in enumerate(bank_txns):
    if i not in matched_bank:
        remain_bank.append({'date': t['date'], 'amount': t['amount'], 'direction': t['direction'], 'idx': i})
remain_ent = []
for i, t in enumerate(ent_txns):
    if i not in matched_ent:
        remain_ent.append({'date': t['date'], 'amount': t['amount'], 'direction': t['direction'], 'idx': i})

print(f"\n=== 剩余: B{len(remain_bank)} + E{len(remain_ent)} = {len(remain_bank)+len(remain_ent)} ===\n")

# ==== STRICT DATE GROUPING ====
# Group by exact date, and by direction pair (CR/借, DR/贷)
from collections import namedtuple

DateGroup = namedtuple('DateGroup', ['date', 'b_items', 'e_items'])

# For remaining items, group by date + direction pair
# Direction: CR→借, DR→贷
groups = []

# Collect all dates that have BOTH bank and enterprise remainders
all_dates = set()
dir_map_remain = {('CR', '借'), ('DR', '贷')}

for (b_dir, e_dir) in dir_map_remain:
    b_by_date = defaultdict(list)
    for t in remain_bank:
        if t['direction'] == b_dir:
            b_by_date[t['date']].append(t)
    e_by_date = defaultdict(list)
    for t in remain_ent:
        if t['direction'] == e_dir:
            e_by_date[t['date']].append(t)
    
    for d in set(b_by_date.keys()) | set(e_by_date.keys()):
        b_items = sorted(b_by_date.get(d, []), key=lambda x: abs(x['amount']))
        e_items = sorted(e_by_date.get(d, []), key=lambda x: abs(x['amount']))
        if b_items or e_items:
            groups.append(DateGroup(d, b_items, e_items))

groups.sort(key=lambda g: g.date)

print(f"严格按日期分组: {len(groups)} 组\n")

# ==== N-LEVEL ANALYSIS ====
print("=" * 90)
print(f"{'日期':<12} {'B':>4} {'E':>4} {'N=2组合':>12} {'N=3组合':>14} {'N=4组合':>16} {'N=5组合':>18} {'N=6组合':>20}")
print("=" * 90)

total_n = {2: 0, 3: 0, 4: 0, 5: 0, 6: 0}
total_n_matches = {2: 0, 3: 0, 4: 0, 5: 0, 6: 0}

for g in groups:
    nB = len(g.b_items)
    nE = len(g.e_items)
    if nB == 0 or nE == 0:
        continue
    
    # Combinations for M:N
    # For 1:N: C(nE, N) per bank item → nB × C(nE, N)
    # For M:1: C(nB, M) per enterprise item → nE × C(nB, M)
    pairs = {}
    for N in [2, 3, 4, 5, 6]:
        # 1 bank = N enterprise
        b1_en = sum(1 for _ in range(nB)) * comb_count(nE, N) if nE >= N else 0
        # M bank = 1 enterprise  
        bm_e1 = sum(1 for _ in range(nE)) * comb_count(nB, N) if nB >= N else 0
        pairs[N] = b1_en + bm_e1
    
    n2 = pairs[2]; n3 = pairs[3]; n4 = pairs[4]; n5 = pairs[5]; n6 = pairs[6]
    
    # Only print non-trivial groups
    total = n2 + n3 + n4 + n5 + n6
    if total > 0:
        print(f"{str(g.date):<12} {nB:>4} {nE:>4} {n2:>12,} {n3:>14,} {n4:>16,} {n5:>18,} {n6:>20,}")
    
    for N in [2, 3, 4, 5, 6]:
        total_n[N] += pairs[N]

print("=" * 90)
print(f"{'合计':<12} {'':>4} {'':>4} {total_n[2]:>12,} {total_n[3]:>14,} {total_n[4]:>16,} {total_n[5]:>18,} {total_n[6]:>20,}")

# ==== ACTUAL MATCH TEST (N≤4 only for feasibility) ====
print("\n" + "=" * 90)
print("实际匹配测试 (N≤4, 严格日期)")
print("=" * 90)

matched_pairs = []  # (b_indices_list, e_indices_list)

for g in groups:
    b_items = g.b_items
    e_items = g.e_items
    
    if not b_items or not e_items:
        continue
    
    # Track which e-items have been used
    used_e = set()
    used_b = set()
    
    for N in [2, 3, 4]:
        # 1 bank = N enterprise
        for bi, bt in enumerate(b_items):
            if bi in used_b: continue
            if N > len(e_items): continue
            for combo in itertools.combinations(range(len(e_items)), N):
                if any(ei in used_e for ei in combo): continue
                e_sum = sum(e_items[ei]['amount'] for ei in combo)
                if abs(bt['amount'] - e_sum) < 0.01:
                    used_b.add(bi)
                    used_e.update(combo)
                    matched_pairs.append(([bt], [e_items[ei] for ei in combo]))
                    break
        
        # N bank = 1 enterprise
        for ei, et in enumerate(e_items):
            if ei in used_e: continue
            if N > len(b_items): continue
            for combo in itertools.combinations(range(len(b_items)), N):
                if any(bi in used_b for bi in combo): continue
                b_sum = sum(b_items[bi]['amount'] for bi in combo)
                if abs(et['amount'] - b_sum) < 0.01:
                    used_e.add(ei)
                    used_b.update(combo)
                    matched_pairs.append(([b_items[bi] for bi in combo], [et]))
                    break
    
    # Move to next group

n2_items = sum(1+N for combo in matched_pairs if len(combo[0])+len(combo[1]) == 3)
n3_items = sum(1+N for combo in matched_pairs if len(combo[0])+len(combo[1]) == 4)
n4_items = sum(1+N for combo in matched_pairs if len(combo[0])+len(combo[1]) >= 5)

print(f"N=2 匹配组: {sum(1 for c in matched_pairs if len(c[0])+len(c[1])==3)} 组, 消掉 {n2_items} 条")
print(f"N=3 匹配组: {sum(1 for c in matched_pairs if len(c[0])+len(c[1])==4)} 组, 消掉 {n3_items} 条")
print(f"N=4 匹配组: {sum(1 for c in matched_pairs if len(c[0])+len(c[1])>=5)} 组, 消掉 {n4_items} 条")
print(f"合计消掉: {n2_items+n3_items+n4_items} 条")
print(f"剩余: {len(remain_bank) + len(remain_ent) - n2_items - n3_items - n4_items} 条")

# Show a few sample matches
print("\n--- 示例匹配 (前5组) ---")
for i, (b_list, e_list) in enumerate(matched_pairs[:5]):
    b_sum = sum(t['amount'] for t in b_list)
    e_sum = sum(t['amount'] for t in e_list)
    print(f"\nMatch {i+1}: {len(b_list)}B <-> {len(e_list)}E  amt={b_sum:.2f}")
    for t in b_list:
        print(f"  B {t['date']} {t['direction']} {t['amount']:>12.2f}")
    for t in e_list:
        print(f"  E {t['date']} {t['direction']} {t['amount']:>12.2f}")
