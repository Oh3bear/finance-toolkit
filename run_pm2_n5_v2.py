"""
±2 窗口 + N≤5 匹配测试。修复版：借方=收入(正)，贷方=支出(负)，
与银行 CR(正)/DR(负) 做符号感知的 1:1 匹配。
跳过 Phase 0，直接 Phase 2 + Phase 2.5 + ±2 N≤5。
"""
import openpyxl, re, math, itertools, time
from datetime import datetime, timedelta
from collections import defaultdict

MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, (int, float)) and 30000 < val < 100000:
        return (datetime(1970, 1, 1) + timedelta(days=val - 25569)).date()
    return None

def parse_amount(s):
    if s is None: return 0.0
    s = str(s).replace(',','').replace(' ','').strip()
    if not s: return 0.0
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    if s.endswith('-'): s = '-' + s[:-1]
    s = s.upper().rstrip('DR').rstrip('CR')
    try: return float(s)
    except: return 0.0

def comb_count(n, k):
    if k > n: return 0
    return math.comb(n, k)

BANK = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX'
ENT  = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX'
TARGET = '01469910120237'

# ==== LOAD BANK ====
wb = openpyxl.load_workbook(BANK, data_only=True)
bank_raw = list(wb.active.iter_rows(values_only=True))
bh = [str(h) if h else '' for h in bank_raw[0]]
acct_col = next(i for i,h in enumerate(bh) if 'Account No' in h)
date_col = next(i for i,h in enumerate(bh) if 'Transaction Date' in h)
drcr_col = next(i for i,h in enumerate(bh) if 'Debit/Credit' in h)
amt_col  = next(i for i,h in enumerate(bh) if 'Amount' in h)
val_col  = next(i for i,h in enumerate(bh) if 'Value Date' in h)
ref_col  = next(i for i,h in enumerate(bh) if 'Customer Reference' in h)

bank_txns = []
for row in bank_raw[1:]:
    acct = str(row[acct_col]).strip().strip("'").strip() if row[acct_col] else ''
    if TARGET not in acct: continue
    raw_date = str(row[date_col]).strip() if row[date_col] else ''
    d = None
    if raw_date:
        m = re.match(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s+(\d{2}):(\d{2})', raw_date)
        if m:
            d = datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1))).date()
    if d is None:
        vd = row[val_col]
        if isinstance(vd, datetime): d = vd.date()
    if d is None: continue
    amt = abs(parse_amount(row[amt_col]))  # Amount always positive
    if amt == 0: continue
    # Direction from Debit/Credit column: values are 'DR' or 'CR'
    dc = str(row[drcr_col]).strip().upper() if row[drcr_col] else ''
    direction = 'DR' if 'DR' in dc else 'CR'
    ref = str(row[ref_col]).strip().strip("'").strip() if row[ref_col] else ''
    bank_txns.append({'date': d, 'amount': amt if direction == 'CR' else -amt, 
                      'direction': direction, 'ref': ref[:60]})

# ==== LOAD ENTERPRISE ====
wb = openpyxl.load_workbook(ENT, data_only=True)
ent_raw = list(wb.active.iter_rows(values_only=True))

ent_txns = []
for row in ent_raw[1:]:
    if not row[12]: continue
    acct = str(row[12]).strip().strip("'").strip()
    if TARGET not in acct: continue
    d = parse_date(row[5])
    if not d: continue
    debit  = parse_amount(row[15]) if len(row) > 15 and row[15] is not None else 0
    credit = parse_amount(row[17]) if len(row) > 17 and row[17] is not None else 0
    # Debit=收入(正), Credit=支出(负)
    if debit != 0:
        ent_txns.append({'date': d, 'amount': debit, 'direction': 'debit'})   # income
    elif credit != 0:
        ent_txns.append({'date': d, 'amount': credit, 'direction': 'credit'})  # expense (negative)

print(f"Total: B{len(bank_txns)} + E{len(ent_txns)}")
print(f"  Bank:   CR={sum(1 for t in bank_txns if t['direction']=='CR')}  DR={sum(1 for t in bank_txns if t['direction']=='DR')}")
print(f"  Ent:    debit(income)={sum(1 for t in ent_txns if t['direction']=='debit')}  credit(expense)={sum(1 for t in ent_txns if t['direction']=='credit')}")

# ==== PHASE 2: 1:1 cents hash, sign-aware, ±7 day ====
# Bank CR(+) matches Ent debit(+); Bank DR(-) matches Ent credit(-)
cents_map = defaultdict(list)
for i, t in enumerate(bank_txns):
    key = (t['direction'], round(abs(t['amount']), 2))
    cents_map[('B', key)].append(i)
for i, t in enumerate(ent_txns):
    e_dir = 'debit' if t['amount'] > 0 else 'credit'
    key = (e_dir, round(abs(t['amount']), 2))
    cents_map[('E', key)].append(i)

# Map bank direction to enterprise direction
dir_pairs = [('CR', 'debit'), ('DR', 'credit')]

mb, me = set(), set()
p2 = 0
for b_dir, e_dir in dir_pairs:
    b_amts = {k[1][1] for k in cents_map if k[0]=='B' and k[1][0]==b_dir}  # k=(src,(dir,amt))
    e_amts = {k[1][1] for k in cents_map if k[0]=='E' and k[1][0]==e_dir}
    for amt in b_amts & e_amts:
        b_idx = cents_map[('B', (b_dir, amt))]
        e_idx = cents_map[('E', (e_dir, amt))]
        for bi in list(b_idx):
            if bi in mb: continue
            for ei in list(e_idx):
                if ei in me: continue
                if abs((bank_txns[bi]['date'] - ent_txns[ei]['date']).days) <= 7:
                    mb.add(bi); me.add(ei)
                    p2 += 2
print(f"Phase2 1:1: {p2} items")

# ==== PHASE 2.5: Same-date bucket by direction ====
date_buckets = defaultdict(lambda: defaultdict(list))
for i, t in enumerate(bank_txns):
    if i in mb: continue
    date_buckets[t['date']]['B_'+t['direction']].append(i)
for i, t in enumerate(ent_txns):
    if i in me: continue
    e_dir = 'debit' if t['amount'] > 0 else 'credit'
    date_buckets[t['date']]['E_'+e_dir].append(i)

p25 = 0
for d, bucket in date_buckets.items():
    for bd, ed in [('B_CR','E_debit'), ('B_DR','E_credit')]:
        bi = bucket.get(bd, [])
        ei = bucket.get(ed, [])
        if not bi or not ei: continue
        b_sum = sum(bank_txns[i]['amount'] for i in bi)
        e_sum = sum(ent_txns[i]['amount'] for i in ei)
        if abs(b_sum - e_sum) < 0.01:
            for i in bi: mb.add(i)
            for i in ei: me.add(i)
            p25 += len(bi) + len(ei)
print(f"Phase2.5 bucket: {p25} items")

# ==== REMAINING ====
remain_b = [t for i,t in enumerate(bank_txns) if i not in mb]
remain_e = [t for i,t in enumerate(ent_txns) if i not in me]
print(f"\nRemaining before Phase3: B{len(remain_b)} + E{len(remain_e)} = {len(remain_b)+len(remain_e)}")

# ==== BUILD ±2 POOLS ====
INDEX = 2
pools = []
for b_dir, e_dir in [('CR','debit'), ('DR','credit')]:
    b_items = [(i, t) for i, t in enumerate(remain_b) if t['direction'] == b_dir]
    e_items = [(i, t) for i, t in enumerate(remain_e) 
               if (e_dir == 'debit' and t['amount'] > 0) or (e_dir == 'credit' and t['amount'] < 0)]
    if not b_items or not e_items: continue
    
    dates = sorted(set(t['date'] for _, t in b_items) | set(t['date'] for _, t in e_items))
    if not dates: continue
    
    current = [dates[0]]
    pool_dates_list = []
    for d in dates[1:]:
        if (d - current[-1]).days <= INDEX:
            current.append(d)
        else:
            pool_dates_list.append(current)
            current = [d]
    pool_dates_list.append(current)
    
    for pdates in pool_dates_list:
        dmin, dmax = min(pdates), max(pdates)
        p_b = [(i, t) for i, t in b_items if dmin <= t['date'] <= dmax]
        p_e = [(i, t) for i, t in e_items if dmin <= t['date'] <= dmax]
        if p_b and p_e:
            pools.append({
                'dates': f"{dmin}~{dmax}" if dmin != dmax else str(dmin),
                'b_items': p_b, 'e_items': p_e,
                'direction': f"{b_dir}/{e_dir}"
            })

print(f"\n+/-2 pools: {len(pools)}")
for p in pools:
    print(f"  {p['dates']:>22} B{len(p['b_items']):>4}+E{len(p['e_items']):>4} {p['direction']}")

# ==== PHASE 3: ±2 window, N≤5 ====
MAX_COMBOS = 10_000_000  # skip C(n,k) > 10M

t0 = time.time()
all_matched_b = set()
all_matched_e = set()
stats = {3: 0, 4: 0, 5: 0, 6: 0}  # total items per match -> count
stats_items = {3: 0, 4: 0, 5: 0, 6: 0}

for pi, pool in enumerate(pools):
    b_items = pool['b_items']
    e_items = pool['e_items']
    nb, ne = len(b_items), len(e_items)
    
    b_amts = [t['amount'] for _, t in b_items]
    e_amts = [t['amount'] for _, t in e_items]
    
    used_b = set()
    used_e = set()
    
    # 1:N (bank -> N enterprise) for N=2..5
    for N in range(2, 6):
        if ne < N: continue
        if comb_count(ne, N) > MAX_COMBOS: continue
        for bi in range(nb):
            if bi in used_b: continue
            for combo in itertools.combinations(range(ne), N):
                if any(ei in used_e for ei in combo): continue
                if abs(b_amts[bi] - sum(e_amts[ei] for ei in combo)) < 0.01:
                    used_b.add(bi)
                    used_e.update(combo)
                    items = 1 + N
                    stats[items] = stats.get(items, 0) + 1
                    stats_items[items] = stats_items.get(items, 0) + items
                    break
    
    # M:1 (M bank -> 1 enterprise) for M=2..5, capped
    for M in range(2, 6):
        if nb < M: continue
        if comb_count(nb, M) > MAX_COMBOS: continue
        for ei in range(ne):
            if ei in used_e: continue
            for combo in itertools.combinations(range(nb), M):
                if any(bi in used_b for bi in combo): continue
                if abs(e_amts[ei] - sum(b_amts[bi] for bi in combo)) < 0.01:
                    used_b.update(combo)
                    used_e.add(ei)
                    items = M + 1
                    stats[items] = stats.get(items, 0) + 1
                    stats_items[items] = stats_items.get(items, 0) + items
                    break
    
    # 2:2, 2:3, 3:2
    for Mb, Ne in [(2,2), (2,3), (3,2)]:
        if nb < Mb or ne < Ne: continue
        cb = comb_count(nb, Mb)
        ce = comb_count(ne, Ne)
        if cb * ce > MAX_COMBOS * 5: continue
        for bc in itertools.combinations(range(nb), Mb):
            if any(bi in used_b for bi in bc): continue
            for ec in itertools.combinations(range(ne), Ne):
                if any(ei in used_e for ei in ec): continue
                if abs(sum(b_amts[bi] for bi in bc) - sum(e_amts[ei] for ei in ec)) < 0.01:
                    used_b.update(bc)
                    used_e.update(ec)
                    items = Mb + Ne
                    stats[items] = stats.get(items, 0) + 1
                    stats_items[items] = stats_items.get(items, 0) + items
                    break
    
    for bi in used_b:
        all_matched_b.add(b_items[bi][0])
    for ei in used_e:
        all_matched_e.add(e_items[ei][0])

elapsed = time.time() - t0

# ==== RESULTS ====
print("\n" + "=" * 65)
print(f"Phase3 (+/-2 window, N<=5) 耗时 {elapsed:.1f}s")
print("=" * 65)
for items in sorted(stats.keys()):
    if stats[items] > 0:
        label = f"{items-1}:1" if items-1 > 1 else "1:1"
        print(f"  1:{items-1} 或 {items-1}:1: {stats[items]} 组, {stats_items[items]} 条")

total_new = sum(stats_items.values())
print(f"  Phase3 总计: {total_new} 条")
print(f"  剩余: B{len(remain_b)-len(all_matched_b)} + E{len(remain_e)-len(all_matched_e)} = "
      f"{len(remain_b)+len(remain_e)-total_new} 条")

total_matched = p2 + p25 + total_new
total_all = len(bank_txns) + len(ent_txns)
print(f"\n  整体匹配率: {total_matched}/{total_all} = {total_matched/total_all*100:.1f}%")
