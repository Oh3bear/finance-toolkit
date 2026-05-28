#!/usr/bin/env python3
"""银企对账 - 1:N 匹配 (N≤3) 分析"""
import openpyxl, re, io, os
from datetime import datetime, timedelta
from collections import defaultdict
from itertools import combinations

OUTPUT = r'C:\Users\CSCI\Downloads\kimiOKC\财务工具集\recon_1to3.txt'
os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)

MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def parse_date(val):
    if isinstance(val, datetime): return val
    if isinstance(val, (int, float)) and 30000 < val < 100000:
        return datetime(1970, 1, 1) + timedelta(days=val - 25569)
    return None

def to_date_key(d):
    return '%d-%02d-%02d' % (d.year, d.month, d.day)

def cents_eq(a, b): return round(a * 100) == round(b * 100)

def within_window(a, b, days=7):
    return abs((a - b).total_seconds()) <= days * 86400

def parse_amount(s):
    if s is None: return 0.0
    s = str(s).replace(',','').replace(' ','').strip()
    if not s: return 0.0
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    if s.endswith('-'): s = '-' + s[:-1]
    s = s.upper().rstrip('DR').rstrip('CR')
    try: return float(s)
    except: return 0.0

BANK = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX'
ENT  = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX'

# ---- read bank ----
wb = openpyxl.load_workbook(BANK, data_only=True)
bank_raw = list(wb.active.iter_rows(values_only=True))
bank_header = [str(h).replace('\n',' ').strip() if h else '' for h in bank_raw[0]]
acct_col = next(i for i,h in enumerate(bank_header) if 'Account No' in h)
date_col = next(i for i,h in enumerate(bank_header) if 'Transaction Date' in h)
drcr_col = next(i for i,h in enumerate(bank_header) if 'Debit/Credit' in h)
amt_col = next(i for i,h in enumerate(bank_header) if 'Amount' in h)
val_col = next(i for i,h in enumerate(bank_header) if 'Value Date' in h)

bank_txns = []
for row in bank_raw[1:]:
    acct = str(row[acct_col]).strip().strip("'").strip() if row[acct_col] else ''
    if not acct: continue
    raw_date = str(row[date_col]).strip() if row[date_col] else ''
    if not raw_date:
        vd = row[val_col]
        if isinstance(vd, datetime): d = vd
        else: continue
    else:
        m = re.match(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s+(\d{2}):(\d{2})', raw_date)
        if m:
            d = datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1)), int(m.group(4)), int(m.group(5)))
        else:
            vd = row[val_col]
            if isinstance(vd, datetime): d = vd
            else: continue
    amt = parse_amount(row[amt_col])
    drcr = str(row[drcr_col]).strip().upper() if row[drcr_col] else ''
    if amt == 0: continue
    direction = 'CR' if drcr == 'CR' else 'DR'
    amount = abs(amt) if drcr == 'CR' else -abs(amt)
    bank_txns.append({'account': acct, 'date': d, 'amount': amount, 'direction': direction})

# ---- read enterprise ----
wb = openpyxl.load_workbook(ENT, data_only=True)
ent_raw = list(wb.active.iter_rows(values_only=True))
ent_header = [str(h).replace('\n',' ').strip() if h else '' for h in ent_raw[0]]

e_acct_col = 12
e_date_col = 5
e_debit_col = 15
e_credit_col = 17

ent_txns = []
for row in ent_raw[1:]:
    if not row[e_acct_col]: continue
    acct = str(row[e_acct_col]).strip().strip("'")
    if not acct: continue
    d = parse_date(row[e_date_col])
    if not d: continue
    debit = parse_amount(row[e_debit_col]) if len(row) > e_debit_col else 0
    credit = parse_amount(row[e_credit_col]) if len(row) > e_credit_col else 0
    if debit != 0:
        ent_txns.append({'account': acct, 'date': d, 'amount': debit, 'direction': 'borrow'})
    elif credit != 0:
        ent_txns.append({'account': acct, 'date': d, 'amount': credit, 'direction': 'lend'})

# ---- account grouping + Phase 0/2/2.5 ----
bank_by_acct = defaultdict(list)
for t in bank_txns: bank_by_acct[t['account']].append(t)
ent_by_acct = defaultdict(list)
for t in ent_txns: ent_by_acct[t['account']].append(t)

total_bank = 0; total_ent = 0; total_quick = 0; total_1to1 = 0; total_bucket = 0
total_rem_b = 0; total_rem_e = 0
total_n2 = 0; total_n3 = 0  # N=2, N=3 match counts

report_lines = []

for acct in sorted(set(bank_by_acct.keys()) & set(ent_by_acct.keys())):
    b_list = sorted(bank_by_acct[acct], key=lambda t: t['date'])
    e_list = sorted(ent_by_acct[acct], key=lambda t: t['date'])
    total_bank += len(b_list); total_ent += len(e_list)

    b_inc = [t for t in b_list if t['direction'] == 'CR']
    b_exp = [t for t in b_list if t['direction'] == 'DR']
    e_deb = [t for t in e_list if t['direction'] == 'borrow']
    e_cre = [t for t in e_list if t['direction'] == 'lend']

    b_inc_sum = sum(t['amount'] for t in b_inc)
    b_exp_sum = sum(t['amount'] for t in b_exp)
    e_deb_sum = sum(t['amount'] for t in e_deb)
    e_cre_sum = sum(t['amount'] for t in e_cre)

    fast_inc = cents_eq(b_inc_sum, e_deb_sum) and b_inc and e_deb
    fast_exp = cents_eq(b_exp_sum, e_cre_sum) and b_exp and e_cre

    used_b = set(); used_e = set()

    if fast_inc:
        for i in range(len(b_list)):
            if b_list[i] in b_inc: used_b.add(i)
        for i in range(len(e_list)):
            if e_list[i] in e_deb: used_e.add(i)
        total_quick += len(b_inc) + len(e_deb)
    if fast_exp:
        for i in range(len(b_list)):
            if b_list[i] in b_exp: used_b.add(i)
        for i in range(len(e_list)):
            if e_list[i] in e_cre: used_e.add(i)
        total_quick += len(b_exp) + len(e_cre)

    # Phase 2: 1:1
    ent_by_cents = defaultdict(list)
    for i, e in enumerate(e_list):
        if i in used_e: continue
        ent_by_cents[round(e['amount'] * 100)].append(i)

    matched_1to1 = 0
    for bi, b in enumerate(b_list):
        if bi in used_b: continue
        c = round(b['amount'] * 100)
        bucket = ent_by_cents.get(c, [])
        if not bucket: continue
        best = None
        for ei in bucket:
            if ei in used_e: continue
            if within_window(b['date'], e_list[ei]['date']):
                best = ei; break
        if best is None:
            for ei in bucket:
                if ei not in used_e:
                    best = ei; break
        if best is not None:
            used_b.add(bi); used_e.add(best); matched_1to1 += 1
    total_1to1 += matched_1to1

    # Phase 2.5
    remaining_b = [(i, b_list[i]) for i in range(len(b_list)) if i not in used_b]
    remaining_e = [(i, e_list[i]) for i in range(len(e_list)) if i not in used_e]

    b_by_date = defaultdict(list)
    for idx, t in remaining_b: b_by_date[to_date_key(t['date'])].append((idx, t))
    e_by_date = defaultdict(list)
    for idx, t in remaining_e: e_by_date[to_date_key(t['date'])].append((idx, t))

    bucket_matched = 0
    for dk, b_items in b_by_date.items():
        e_items = e_by_date.get(dk, [])
        b_inc_d = [(i,t) for i,t in b_items if t['amount'] > 0]
        e_deb_d = [(i,t) for i,t in e_items if t['amount'] > 0 and t['direction'] == 'borrow']
        if b_inc_d and e_deb_d:
            bs = sum(t['amount'] for _,t in b_inc_d)
            es = sum(t['amount'] for _,t in e_deb_d)
            if cents_eq(bs, es):
                for i,_ in b_inc_d: used_b.add(i)
                for i,_ in e_deb_d: used_e.add(i)
                bucket_matched += len(b_inc_d) + len(e_deb_d)
        b_exp_d = [(i,t) for i,t in b_items if t['amount'] < 0]
        e_cre_d = [(i,t) for i,t in e_items if t['amount'] < 0 and t['direction'] == 'lend']
        if b_exp_d and e_cre_d:
            bs = sum(t['amount'] for _,t in b_exp_d)
            es = sum(t['amount'] for _,t in e_cre_d)
            if cents_eq(bs, es):
                for i,_ in b_exp_d: used_b.add(i)
                for i,_ in e_cre_d: used_e.add(i)
                bucket_matched += len(b_exp_d) + len(e_cre_d)
    total_bucket += bucket_matched

    # ---- Phase 3: 1:N (N≤3) matching ----
    rem_b = [(i, b_list[i]) for i in range(len(b_list)) if i not in used_b]
    rem_e = [(i, e_list[i]) for i in range(len(e_list)) if i not in used_e]

    b_rem_count = len(rem_b)
    e_rem_count = len(rem_e)

    n2_matched = 0
    n3_matched = 0
    detail = []

    # Helper: find 1:N matches (N=2 or 3) in one direction
    def find_1toN(source, target, n):
        """source: list of (idx, txn), target: list of (idx, txn). Try to match ONE source to N targets."""
        matched_source = set()
        used_target_idxs = set()
        groups = []

        for si, s_txn in source:
            if si in matched_source: continue
            s_amt = s_txn['amount']

            # Filter targets by date window and direction
            valid_targets = [
                (ti, t_txn) for ti, t_txn in target
                if ti not in used_target_idxs
                and within_window(s_txn['date'], t_txn['date'])
            ]

            # Try N=2 combinations
            if n >= 2 and len(valid_targets) >= 2:
                for (t1i, t1), (t2i, t2) in combinations(valid_targets, 2):
                    if cents_eq(s_amt, t1['amount'] + t2['amount']):
                        matched_source.add(si)
                        used_target_idxs.add(t1i)
                        used_target_idxs.add(t2i)
                        groups.append((si, s_txn, [(t1i, t1), (t2i, t2)]))
                        break
                if si in matched_source: continue

            # Try N=3 combinations
            if n >= 3 and len(valid_targets) >= 3:
                for (t1i, t1), (t2i, t2), (t3i, t3) in combinations(valid_targets, 3):
                    if cents_eq(s_amt, t1['amount'] + t2['amount'] + t3['amount']):
                        matched_source.add(si)
                        used_target_idxs.add(t1i)
                        used_target_idxs.add(t2i)
                        used_target_idxs.add(t3i)
                        groups.append((si, s_txn, [(t1i, t1), (t2i, t2), (t3i, t3)]))
                        break

        return matched_source, used_target_idxs, groups

    # Match by direction groups
    # Bank CR (positive) ↔ Enterprise borrow (positive)
    b_inc_rem = [(i, t) for i, t in rem_b if t['amount'] > 0]
    e_deb_rem = [(i, t) for i, t in rem_e if t['amount'] > 0 and t['direction'] == 'borrow']

    # Bank DR (negative) ↔ Enterprise lend (negative)
    b_exp_rem = [(i, t) for i, t in rem_b if t['amount'] < 0]
    e_cre_rem = [(i, t) for i, t in rem_e if t['amount'] < 0 and t['direction'] == 'lend']

    all_match_groups = []

    for src_name, src, tgt, tgt_name in [
        ('B→E(收入)', b_inc_rem, e_deb_rem, 'E(借)'),
        ('E→B(收入)', e_deb_rem, b_inc_rem, 'B(贷)'),
        ('B→E(支出)', b_exp_rem, e_cre_rem, 'E(贷)'),
        ('E→B(支出)', e_cre_rem, b_exp_rem, 'B(借)'),
    ]:
        m_src, m_tgt, groups = find_1toN(src, tgt, 3)
        for si, s_txn, targets in groups:
            n = len(targets)
            all_match_groups.append((acct, src_name, si, s_txn, targets, n))

        # Mark used
        for si in m_src:
            used_b.add(si) if src == b_inc_rem or src == b_exp_rem else used_e.add(si)
        for ti in m_tgt:
            used_e.add(ti) if tgt == e_deb_rem or tgt == e_cre_rem else used_b.add(ti)

        for entry in all_match_groups[-len(groups):]:
            n = entry[5]
            if n == 2: n2_matched += 1 + n
            else: n3_matched += 1 + n

    rem_b_after = len([i for i in range(len(b_list)) if i not in used_b])
    rem_e_after = len([i for i in range(len(e_list)) if i not in used_e])

    if b_rem_count > 0 or e_rem_count > 0:
        report_lines.append(
            '\n%s: 进入Phase3 B%d+E%d → 1:N后剩余 B%d+E%d (N=2:%d条, N=3:%d条)' % (
                acct, b_rem_count, e_rem_count, rem_b_after, rem_e_after,
                n2_matched, n3_matched))

    total_n2 += n2_matched
    total_n3 += n3_matched
    total_rem_b += rem_b_after
    total_rem_e += rem_e_after

# ---- Output ----
out = io.StringIO()

out.write('===== 1:N (N≤3) 匹配分析 =====\n\n')
out.write('Phase0 快速通道: %d 条 (%d 对)\n' % (total_quick, total_quick // 2))
out.write('Phase2 1:1 匹配:  %d 对\n' % total_1to1)
out.write('Phase2.5 日期桶: %d 条\n' % total_bucket)

matched_before = total_bank + total_ent - sum(
    int(l.split('进入Phase3 B')[1].split('+')[0]) if '进入Phase3' in l else 0
    for l in report_lines) - sum(
    int(l.split('+E')[1].split(' ')[0]) if '进入Phase3' in l else 0
    for l in report_lines)

out.write('Phase3前剩余: B%d + E%d\n' % (
    sum(int(l.split('进入Phase3 B')[1].split('+')[0]) for l in report_lines if '进入Phase3' in l),
    sum(int(l.split('+E')[1].split(' ')[0]) for l in report_lines if '进入Phase3' in l)))

out.write('\n--- Phase3 1:N 结果 ---\n')
out.write('N=2 消掉: %d 条\n' % total_n2)
out.write('N=3 消掉: %d 条\n' % total_n3)

out.write('\n--- 各账户详情 ---\n')
for l in report_lines:
    out.write(l + '\n')

matched_after = total_bank + total_ent - total_rem_b - total_rem_e
rate = matched_after / (total_bank + total_ent) * 100 if (total_bank + total_ent) > 0 else 0

out.write('\n===== 最终汇总 =====\n')
out.write('总交易: Bank %d + Ent %d = %d\n' % (total_bank, total_ent, total_bank + total_ent))
out.write('Phase0: %d | Phase2: %d | Phase2.5: %d | Phase3 N=2: %d | Phase3 N=3: %d\n' % (
    total_quick, total_1to1 * 2, total_bucket, total_n2, total_n3))
out.write('已匹配: %d / %d = %.1f%%\n' % (matched_after, total_bank + total_ent, rate))
out.write('剩余: B%d + E%d = %d\n' % (total_rem_b, total_rem_e, total_rem_b + total_rem_e))

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(out.getvalue())
print(out.getvalue())
