#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从小到大匹配：剩余条目按金额升序，逐步消费，验证有效性"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl, re, math, itertools
from datetime import datetime, timedelta
from collections import defaultdict

# ── 配置 ──
TARGET = '01469910120237'
WINDOW_DAYS = 2
MAX_DEPTH = 15
MAX_COMBOS = 5_000_000  # per item, safety limit
TOLERANCE = 0.02

# ── 日期/金额解析 ──
MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, (int, float)) and 30000 < val < 100000:
        return (datetime(1970, 1, 1) + timedelta(days=val - 25569)).date()
    return None

def parse_amount(s):
    if s is None: return 0.0
    s = str(s).replace(',','').replace(' ','').strip()
    if not s: return 0.0
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    if s.endswith('-'): s = '-' + s[:-1]
    s = s.upper().rstrip('DR').rstrip('CR')
    try: return float(s)
    except: return 0.0

# ── 读取数据 ──
BANK = r"C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX"
ENT  = r"C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX"

# Bank
wb = openpyxl.load_workbook(BANK, data_only=True)
bank_raw = list(wb.active.iter_rows(values_only=True))
bh = [str(h) if h else '' for h in bank_raw[0]]
acct_col = next(i for i,h in enumerate(bh) if 'Account No' in h)
date_col = next(i for i,h in enumerate(bh) if 'Transaction Date' in h)
drcr_col = next(i for i,h in enumerate(bh) if 'Debit/Credit' in h)
amt_col  = next(i for i,h in enumerate(bh) if 'Amount' in h)
val_col  = next(i for i,h in enumerate(bh) if 'Value Date' in h)
ref_col  = next(i for i,h in enumerate(bh) if 'Customer Reference' in h)

bank_txns = []
for row in bank_raw[1:]:
    acct = str(row[acct_col]).strip().strip("'").strip() if row[acct_col] else ''
    if TARGET not in acct: continue
    raw_date = str(row[date_col]).strip() if row[date_col] else ''
    d = None
    if raw_date:
        m = re.match(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s+(\d{2}):(\d{2})', raw_date)
        if m:
            d = datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1))).date()
    if d is None:
        vd = row[val_col]
        if isinstance(vd, datetime): d = vd.date()
    if d is None: continue
    amt = abs(parse_amount(row[amt_col]))
    if amt == 0: continue
    dc = str(row[drcr_col]).strip().upper() if row[drcr_col] else ''
    direction = 'DR' if 'DR' in dc else 'CR'
    ref = str(row[ref_col]).strip().strip("'").strip() if row[ref_col] else ''
    bank_txns.append({'date': d, 'amount': amt if direction == 'CR' else -amt, 
                      'direction': direction, 'ref': ref[:60]})

# Enterprise
wb = openpyxl.load_workbook(ENT, data_only=True)
ent_raw = list(wb.active.iter_rows(values_only=True))
e_date_col = 5; e_acct_col = 12; e_debit_col = 15; e_credit_col = 17; e_memo_col = 11

ent_txns = []
for row in ent_raw[1:]:
    if not row[e_acct_col]: continue
    acct = str(row[e_acct_col]).strip().strip("'").strip()
    if TARGET not in acct: continue
    d = parse_date(row[e_date_col])
    if not d: continue
    debit  = parse_amount(row[e_debit_col]) if len(row) > e_debit_col and row[e_debit_col] is not None else 0
    credit = parse_amount(row[e_credit_col]) if len(row) > e_credit_col and row[e_credit_col] is not None else 0
    memo = str(row[e_memo_col]).strip()[:60] if row[e_memo_col] else ''
    if debit != 0:
        ent_txns.append({'date': d, 'amount': debit, 'direction': 'debit', 'memo': memo})
    elif credit != 0:
        ent_txns.append({'date': d, 'amount': credit, 'direction': 'credit', 'memo': memo})

print(f"Bank: {len(bank_txns)}  Enterprise: {len(ent_txns)}")

# ── Phase 2: 1:1 (±7 days) ──
cents_map = defaultdict(list)
for i, t in enumerate(bank_txns):
    cents_map[('B', t['direction'], round(abs(t['amount']), 2))].append(i)
for i, t in enumerate(ent_txns):
    e_dir = 'debit' if t['amount'] > 0 else 'credit'
    cents_map[('E', e_dir, round(abs(t['amount']), 2))].append(i)

matched_b = set(); matched_e = set()
p2 = 0
for b_dir, e_dir in [('CR','debit'), ('DR','credit')]:
    b_amts = {k[2] for k in cents_map if k[0]=='B' and k[1]==b_dir}
    e_amts = {k[2] for k in cents_map if k[0]=='E' and k[1]==e_dir}
    for amt in b_amts & e_amts:
        for bi in list(cents_map[('B', b_dir, amt)]):
            if bi in matched_b: continue
            for ei in list(cents_map[('E', e_dir, amt)]):
                if ei in matched_e: continue
                if abs((bank_txns[bi]['date'] - ent_txns[ei]['date']).days) <= 7:
                    matched_b.add(bi); matched_e.add(ei)
                    p2 += 2
print(f"Phase2 1:1: {p2} items ({p2//2} pairs)")

# ── Phase 2.5: same-date bucket ──
date_buckets = defaultdict(lambda: defaultdict(list))
for i, t in enumerate(bank_txns):
    if i in matched_b: continue
    date_buckets[t['date']]['B_'+t['direction']].append(i)
for i, t in enumerate(ent_txns):
    if i in matched_e: continue
    date_buckets[t['date']]['E_debit' if t['amount']>0 else 'E_credit'].append(i)

p25 = 0
for d, bucket in date_buckets.items():
    for bd, ed in [('B_CR','E_debit'), ('B_DR','E_credit')]:
        bi = bucket.get(bd, []); ei = bucket.get(ed, [])
        if bi and ei and len(bi)+len(ei) > 1:
            if abs(sum(bank_txns[i]['amount'] for i in bi) - 
                   sum(ent_txns[i]['amount'] for i in ei)) < 0.01:
                for i in bi: matched_b.add(i)
                for i in ei: matched_e.add(i)
                p25 += len(bi) + len(ei)
print(f"Phase2.5 bucket: {p25} items")

# ── 剩余条目 ──
remain_b = [(i, t) for i, t in enumerate(bank_txns) if i not in matched_b]
remain_e = [(i, t) for i, t in enumerate(ent_txns) if i not in matched_e]
print(f"Remaining: B{len(remain_b)} E{len(remain_e)}")

# ── 按 ±2 窗口分组 ──
INDEX = WINDOW_DAYS
pools = []
for b_dir in ['CR', 'DR']:
    b_items = [(orig_i, t) for orig_i, t in remain_b if t['direction'] == b_dir]
    e_dir = 'debit' if b_dir == 'CR' else 'credit'
    e_items = [(orig_i, t) for orig_i, t in remain_e if 
               (e_dir == 'debit' and t['amount'] > 0) or 
               (e_dir == 'credit' and t['amount'] < 0)]
    if not b_items or not e_items: continue
    
    dates = sorted(set(t['date'] for _, t in b_items) | set(t['date'] for _, t in e_items))
    current = [dates[0]]
    pool_dates = []
    for d in dates[1:]:
        if (d - current[-1]).days <= INDEX:
            current.append(d)
        else:
            pool_dates.append(current)
            current = [d]
    pool_dates.append(current)
    
    for pdates in pool_dates:
        dmin, dmax = min(pdates), max(pdates)
        p_b = [(orig_i, t) for orig_i, t in b_items if dmin <= t['date'] <= dmax]
        p_e = [(orig_i, t) for orig_i, t in e_items if dmin <= t['date'] <= dmax]
        if p_b and p_e:
            pools.append({
                'dates': f"{dmin}~{dmax}" if dmin != dmax else str(dmin),
                'b_items': p_b, 'e_items': p_e,
                'b_dir': b_dir
            })

print(f"\n{len(pools)} pools after ±{INDEX}d grouping:")
for p in pools:
    print(f"  {p['dates']:20s}  B{p['b_dir']}:{len(p['b_items']):3d}  E:{len(p['e_items']):3d}")

# ── 子集查找：从小到大消费 ──
def find_subset(target, candidates, max_depth, max_combos):
    """在 candidates 中找和为 target 的子集（绝对值匹配）。
    candidates: [(orig_idx, txn_dict)], 已按 |amount| 升序
    返回: (matched_indices, depth) 或 (None, reason)"""
    target_abs = abs(target)
    
    # 过滤：只用金额 ≤ target_abs 的条目（绝对值）
    valid = [(idx, txn) for idx, txn in candidates if abs(txn['amount']) <= target_abs + TOLERANCE]
    if not valid:
        return (None, "no candidates <= target")
    
    # 如果恰好一条对得上
    for idx, txn in valid:
        if abs(abs(txn['amount']) - target_abs) < TOLERANCE:
            return ([idx], 1)
    
    total_combo = 0
    # 按 N 递增
    for depth in range(2, min(max_depth + 1, len(valid) + 1)):
        n_combos = math.comb(len(valid), depth)
        if n_combos > max_combos:
            return (None, f"N={depth} combos {n_combos:,} > limit {max_combos:,}")
        
        for combo in itertools.combinations(valid, depth):
            total_combo += 1
            s = sum(abs(txn['amount']) for _, txn in combo)
            if abs(s - target_abs) < TOLERANCE:
                return ([idx for idx, _ in combo], depth)
            
            if total_combo >= max_combos:
                return (None, f"exhausted {max_combos:,} combos at N={depth}")
    
    # Depth exceeded
    if len(valid) > max_depth:
        return (None, f"need N>{max_depth} (valid={len(valid)})")
    return (None, "no subset found")

# ── 逐池处理 ──
total_p3 = 0
unmatched_stats = []

for pool in pools:
    b_dict = {orig_i: t for orig_i, t in pool['b_items']}
    e_dict = {orig_i: t for orig_i, t in pool['e_items']}
    
    # 银行条目按绝对值升序
    b_sorted = sorted(pool['b_items'], key=lambda x: abs(x[1]['amount']))
    
    consumed_b = set()  # 已消费的银行条目
    consumed_e = set()  # 已消费的企业条目
    pool_matched_b = 0
    pool_matched_e_count = 0
    pool_match_groups = []
    
    for b_orig, b_txn in b_sorted:
        if b_orig in consumed_b: continue
        target = b_txn['amount']
        target_abs = abs(target)
        
        # 剩余企业条目，按绝对值升序
        remaining_e = [(orig_i, t) for orig_i, t in pool['e_items'] 
                       if orig_i not in consumed_e]
        e_sorted = sorted(remaining_e, key=lambda x: abs(x[1]['amount']))
        
        if not e_sorted: break
        
        result = find_subset(target, e_sorted, MAX_DEPTH, MAX_COMBOS)
        
        if result[0] is not None:
            matched_indices = result[0]
            depth = result[1]
            consumed_b.add(b_orig)
            for ei in matched_indices:
                consumed_e.add(ei)
            pool_matched_b += 1
            pool_matched_e_count += len(matched_indices)
            pool_match_groups.append((b_orig, matched_indices, depth, target))
        else:
            # 没找到 —— 也试试反向：企业条目凑银行
            pass
    
    # 反向：未被消费的企业条目，尝试凑剩余银行
    b_remaining = [orig_i for orig_i, _ in b_sorted if orig_i not in consumed_b]
    e_unconsumed = [orig_i for orig_i, _ in pool['e_items'] if orig_i not in consumed_e]

    if e_unconsumed and b_remaining:
        e_dict_local = {orig_i: t for orig_i, t in pool['e_items']}
        e_sorted_rev = sorted([(orig_i, e_dict_local[orig_i]) for orig_i in e_unconsumed],
                              key=lambda x: abs(x[1]['amount']))
        
        for e_orig, e_txn in e_sorted_rev:
            if e_orig in consumed_e: continue
            target = -e_txn['amount']  # opposite sign for match
            
            # Rebuild remaining bank items
            b_dict_local = {orig_i: bank_txns[orig_i] for orig_i in b_remaining
                          if orig_i not in consumed_b}
            current_b = [(orig_i, b_dict_local[orig_i]) for orig_i in b_dict_local]
            b_sorted_rev = sorted(current_b, key=lambda x: abs(x[1]['amount']))
            
            if not b_sorted_rev: break
            
            result = find_subset(target, b_sorted_rev, MAX_DEPTH, MAX_COMBOS)
            if result[0] is not None:
                matched_indices = result[0]
                depth = result[1]
                consumed_e.add(e_orig)
                for bi in matched_indices:
                    consumed_b.add(bi)
                pool_matched_b += len(matched_indices)
                pool_matched_e_count += 1
                pool_match_groups.append((e_orig, matched_indices, depth, target, True))
    
    total_p3 += pool_matched_b + pool_matched_e_count
    
    # 剩余统计
    b_unmatched = len([orig_i for orig_i, _ in pool['b_items'] if orig_i not in consumed_b])
    e_unmatched_final = len([orig_i for orig_i, _ in pool['e_items'] 
                             if orig_i not in consumed_e])
    
    print(f"\nPool {pool['dates']} {pool['b_dir']}:")
    print(f"  B{len(pool['b_items'])} E{len(pool['e_items'])} -> matched B{pool_matched_b} E{pool_matched_e_count}")
    print(f"  remaining B{b_unmatched} E{e_unmatched_final}")
    for g in pool_match_groups[:10]:
        is_rev = len(g) >= 6
        if is_rev:
            print(f"  {len(g[1])}B = 1E(CNY{abs(g[3]):,.2f})  depth={g[2]}")
        else:
            print(f"  1B(CNY{abs(g[3]):,.2f}) = {len(g[1])}E  depth={g[2]}")
    if len(pool_match_groups) > 10:
        print(f"  ... and {len(pool_match_groups)-10} more groups")

print(f"\n{'='*50}")
print(f"Phase3 small-first total: {total_p3} items")
print(f"Overall: Phase2({p2}) + Phase2.5({p25}) + Phase3({total_p3}) = {p2+p25+total_p3}")
print(f"Match rate: {(p2+p25+total_p3)/(len(bank_txns)+len(ent_txns))*100:.1f}%")
