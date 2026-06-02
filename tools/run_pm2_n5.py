"""
±2 窗口 + N≤5 实际匹配测试。
策略：候选池按±2天合并，在每个池内搜M:N（M+N≤5），
对大边（>30条）只做 1:N 方向，跳过 M:1 避免 C(n,5) 爆炸。
"""
import openpyxl, re, math, itertools, time
from datetime import datetime, timedelta
from collections import defaultdict

MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

def parse_date(val):
    if isinstance(val, datetime): return val.date()
    if isinstance(val, (int, float)) and 30000 < val < 100000:
        return (datetime(1970, 1, 1) + timedelta(days=val - 25569)).date()
    return None

def parse_amount(s):
    if s is None: return 0.0
    s = str(s).replace(',','').replace(' ','').strip()
    if not s: return 0.0
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    if s.endswith('-'): s = '-' + s[:-1]
    s = s.upper().rstrip('DR').rstrip('CR')
    try: return float(s)
    except: return 0.0

def comb_count(n, k):
    if k > n: return 0
    return math.comb(n, k)

# ==== LOAD ====
BANK = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX'
ENT  = r'C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX'
TARGET = '01469910120237'

wb = openpyxl.load_workbook(BANK, data_only=True)
bank_raw = list(wb.active.iter_rows(values_only=True))
bh = [str(h) if h else '' for h in bank_raw[0]]
acct_col = next(i for i,h in enumerate(bh) if 'Account No' in h)
date_col = next(i for i,h in enumerate(bh) if 'Transaction Date' in h)
drcr_col = next(i for i,h in enumerate(bh) if 'Debit/Credit' in h)
amt_col = next(i for i,h in enumerate(bh) if 'Amount' in h)
val_col = next(i for i,h in enumerate(bh) if 'Value Date' in h)
ref_col = next(i for i,h in enumerate(bh) if 'Customer Reference' in h)

bank_txns = []
for row in bank_raw[1:]:
    acct = str(row[acct_col]).strip().strip("'").strip() if row[acct_col] else ''
    if TARGET not in acct: continue
    raw_date = str(row[date_col]).strip() if row[date_col] else ''
    d = None
    if raw_date:
        m = re.match(r'(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})\s+(\d{2}):(\d{2})', raw_date)
        if m:
            d = datetime(int(m.group(3)), MONTHS[m.group(2).lower()], int(m.group(1))).date()
    if d is None:
        vd = row[val_col]
        if isinstance(vd, datetime): d = vd.date()
    if d is None: continue
    amt = parse_amount(row[amt_col])
    if amt == 0: continue
    ref = str(row[ref_col]).strip().strip("'").strip() if row[ref_col] else ''
    direction = 'CR' if amt > 0 else 'DR'
    bank_txns.append({'date': d, 'amount': amt, 'direction': direction, 'ref': ref[:60]})

wb = openpyxl.load_workbook(ENT, data_only=True)
ent_raw = list(wb.active.iter_rows(values_only=True))
e_date_col, e_acct_col, e_debit_col, e_credit_col = 5, 12, 15, 17

ent_txns = []
for row in ent_raw[1:]:
    if not row[e_acct_col]: continue
    acct = str(row[e_acct_col]).strip().strip("'").strip()
    if TARGET not in acct: continue
    d = parse_date(row[e_date_col])
    if not d: continue
    debit = parse_amount(row[e_debit_col])
    credit = parse_amount(row[e_credit_col])
    # SAP: 借方=正收入, 贷方=负支出(预签负数), 统一放入 'lend' 方向
    if debit != 0:
        ent_txns.append({'date': d, 'amount': debit, 'direction': 'lend'})
    elif credit != 0:
        ent_txns.append({'date': d, 'amount': credit, 'direction': 'lend'})

print(f"Total: B{len(bank_txns)} + E{len(ent_txns)}")

# ==== PHASE 0: Account-level sum check ====
# Bank CR(收入) vs Enterprise lend(收入=正+支出=负) 
# Bank DR(支出) vs Enterprise lend (same pool, different sign)
b_cr_set = {i for i,t in enumerate(bank_txns) if t['direction'] == 'CR'}
b_dr_set = {i for i,t in enumerate(bank_txns) if t['direction'] == 'DR'}
e_all_set = {i for i,t in enumerate(ent_txns)}  # all enterprise items

b_cr_sum = sum(bank_txns[i]['amount'] for i in b_cr_set)
b_dr_sum = sum(bank_txns[i]['amount'] for i in b_dr_set)
e_all_sum = sum(ent_txns[i]['amount'] for i in e_all_set)

mb, me = set(), set()
p0 = 0
if abs(b_cr_sum - e_all_sum) < 0.01:
    mb.update(b_cr_set); me.update(e_all_set)
    p0 += len(b_cr_set) + len(e_all_set)
    print(f"Phase0 CR: B{len(b_cr_set)}+E{len(e_all_set)}")
elif abs(b_dr_sum - e_all_sum) < 0.01:
    mb.update(b_dr_set); me.update(e_all_set)
    p0 += len(b_dr_set) + len(e_all_set)
    print(f"Phase0 DR: B{len(b_dr_set)}+E{len(e_all_set)}")

# ==== PHASE 2: 1:1 ±7 days ====
cents_map = defaultdict(list)
for i, t in enumerate(bank_txns):
    if i in mb: continue
    cents_map[('B', t['direction'], round(abs(t['amount']), 2))].append(i)
for i, t in enumerate(ent_txns):
    if i in me: continue
    # All enterprise items are 'lend' direction; match abs amount
    cents_map[('E', 'lend', round(abs(t['amount']), 2))].append(i)

p2 = 0
for b_dir in ['CR', 'DR']:
    b_amts = {k[2] for k in cents_map if k[0]=='B' and k[1]==b_dir}
    e_amts = {k[2] for k in cents_map if k[0]=='E' and k[1]=='lend'}
    for amt in b_amts & e_amts:
        b_idx = cents_map[('B', b_dir, amt)]
        e_idx = cents_map[('E', 'lend', amt)]
        for bi in list(b_idx):
            if bi in mb: continue
            bamt = bank_txns[bi]['amount']
            for ei in list(e_idx):
                if ei in me: continue
                eamt = ent_txns[ei]['amount']
                # Must have same sign (income vs income, expense vs expense)
                same_sign = (bamt > 0 and eamt > 0) or (bamt < 0 and eamt < 0)
                if not same_sign: continue
                if abs((bank_txns[bi]['date'] - ent_txns[ei]['date']).days) <= 7:
                    mb.add(bi); me.add(ei)
                    p2 += 2
print(f"Phase2 1:1: {p2} items")

# ==== PHASE 2.5: same-date bucket ====
date_buckets = defaultdict(lambda: defaultdict(list))
for i, t in enumerate(bank_txns):
    if i in mb: continue
    date_buckets[t['date']]['B_'+t['direction']].append(i)
for i, t in enumerate(ent_txns):
    if i in me: continue
    date_buckets[t['date']]['E_lend'].append(i)

p25 = 0
for d, bucket in date_buckets.items():
    for bd in ['B_CR', 'B_DR']:
        bi = bucket.get(bd, [])
        ei = bucket.get('E_lend', [])
        if not bi or not ei: continue
        # Match if same-date sum matches
        b_sum = sum(bank_txns[i]['amount'] for i in bi)
        e_sum = sum(ent_txns[i]['amount'] for i in ei)
        if abs(b_sum - e_sum) < 0.01:
            for i in bi: mb.add(i)
            for i in ei: me.add(i)
            p25 += len(bi) + len(ei)
print(f"Phase2.5 bucket: {p25} items")

# ==== REMAINING ====
remain_b = [t for i,t in enumerate(bank_txns) if i not in mb]
remain_e = [t for i,t in enumerate(ent_txns) if i not in me]
print(f"\nBefore Phase3: B{len(remain_b)} + E{len(remain_e)} = {len(remain_b)+len(remain_e)}")

# ==== BUILD ±2 POOLS ====
INDEX = 2
all_dates = sorted(set(t['date'] for t in remain_b) | set(t['date'] for t in remain_e))

pools = []
for b_dir in ['CR', 'DR']:
    b_items = [(i, t) for i, t in enumerate(remain_b) if t['direction'] == b_dir]
    e_items = [(i, t) for i, t in enumerate(remain_e)]  # all enterprise items are 'lend'
    if not b_items or not e_items:
        continue
    dates = sorted(set(t['date'] for _, t in b_items) | set(t['date'] for _, t in e_items))
    if not dates:
        continue
    current = [dates[0]]
    pool_dates = []
    for d in dates[1:]:
        if (d - current[-1]).days <= INDEX:
            current.append(d)
        else:
            pool_dates.append(current)
            current = [d]
    pool_dates.append(current)
    
    for pdates in pool_dates:
        dmin, dmax = min(pdates), max(pdates)
        p_b = [(i, t) for i, t in b_items if dmin <= t['date'] <= dmax]
        p_e = [(i, t) for i, t in e_items if dmin <= t['date'] <= dmax]
        if p_b and p_e:
            pools.append({
                'dates': f"{dmin}~{dmax}" if dmin != dmax else str(dmin),
                'b_items': p_b,
                'e_items': p_e,
                'direction': f"{b_dir}/lend"
            })

print(f"\n±2 pools: {len(pools)}")
for p in pools:
    print(f"  {p['dates']:>22}  B{len(p['b_items']):>3}+E{len(p['e_items']):>3}  {p['direction']}")

# ==== PHASE 3: M:N with N≤5 ====
MAX_ONE_SIDE_COMBOS = 5_000_000  # Skip C(n,k) > 5M

stats = {'n2': 0, 'n3': 0, 'n4': 0, 'n5': 0, 'n2_items': 0, 'n3_items': 0, 'n4_items': 0, 'n5_items': 0}
all_matched_b = set()  # indices into remain_b
all_matched_e = set()  # indices into remain_e

t0 = time.time()

for pi, pool in enumerate(pools):
    b_items = pool['b_items']  # (idx, tx)
    e_items = pool['e_items']
    
    nb, ne = len(b_items), len(e_items)
    used_b = set()
    used_e = set()
    
    # Build amount arrays for fast lookup
    b_amts = [t['amount'] for _, t in b_items]
    e_amts = [t['amount'] for _, t in e_items]
    
    def try_match(mask_b, mask_e, label):
        """Try to match b[mask_b] sum == e[mask_e] sum"""
        b_sum = sum(b_amts[bi] for bi in mask_b)
        e_sum = sum(e_amts[ei] for ei in mask_e)
        if abs(b_sum - e_sum) < 0.01:
            # Check none are used
            if any(bi in used_b for bi in mask_b) or any(ei in used_e for ei in mask_e):
                return False
            used_b.update(mask_b)
            used_e.update(mask_e)
            return True
        return False
    
    pool_matches = []
    
    # Try 1:N (N=2..5)
    for N in [2, 3, 4, 5]:
        if ne < N: continue
        combos_n = comb_count(ne, N)
        if combos_n > MAX_ONE_SIDE_COMBOS: continue
        for bi in range(nb):
            if bi in used_b: continue
            for combo in itertools.combinations(range(ne), N):
                if any(ei in used_e for ei in combo): continue
                if try_match([bi], list(combo), f"1:{N}"):
                    pool_matches.append(([bi], list(combo)))
                    break
    
    # Try M:1 (M=2..3, capped for large B)
    for M in [2, 3]:
        if nb < M: continue
        combos_m = comb_count(nb, M)
        if combos_m > MAX_ONE_SIDE_COMBOS: continue
        for ei in range(ne):
            if ei in used_e: continue
            for combo in itertools.combinations(range(nb), M):
                if any(bi in used_b for bi in combo): continue
                if try_match(list(combo), [ei], f"{M}:1"):
                    pool_matches.append((list(combo), [ei]))
                    break
    
    # Try 2:2
    if nb >= 2 and ne >= 2:
        cb2 = comb_count(nb, 2)
        ce2 = comb_count(ne, 2)
        if cb2 * ce2 <= MAX_ONE_SIDE_COMBOS * 10:  # allow larger for 2:2
            for bi_combo in itertools.combinations(range(nb), 2):
                if any(bi in used_b for bi in bi_combo): continue
                for ei_combo in itertools.combinations(range(ne), 2):
                    if any(ei in used_e for ei in ei_combo): continue
                    if try_match(list(bi_combo), list(ei_combo), "2:2"):
                        pool_matches.append((list(bi_combo), list(ei_combo)))
                        break
    
    # Count
    for b_mask, e_mask in pool_matches:
        total_items = len(b_mask) + len(e_mask)
        if total_items == 3:
            stats['n2'] += 1
            stats['n2_items'] += total_items
        elif total_items == 4:
            stats['n3'] += 1
            stats['n3_items'] += total_items
        elif total_items <= 5:
            stats['n4'] += 1
            stats['n4_items'] += total_items
        elif total_items <= 6:
            stats['n5'] += 1
            stats['n5_items'] += total_items
    
    for bi in used_b:
        all_matched_b.add(b_items[bi][0])
    for ei in used_e:
        all_matched_e.add(e_items[ei][0])
    
    if pool_matches:
        print(f"\nPool {pi}: {pool['dates']} B{nb}+E{ne} → {len(pool_matches)} groups")
        for bm, em in pool_matches:
            b_sum = sum(b_amts[bi] for bi in bm)
            e_sum = sum(e_amts[ei] for ei in em)
            print(f"  {len(bm)}B<->{len(em)}E  amt={b_sum:,.2f}")
            for bi in bm:
                t = b_items[bi][1]
                print(f"    B {t['date']} {t['direction']} {t['amount']:>15,.2f}")
            for ei in em:
                t = e_items[ei][1]
                print(f"    E {t['date']} {t['direction']} {t['amount']:>15,.2f}")

elapsed = time.time() - t0

# ==== RESULTS ====
print("\n" + "=" * 70)
print(f"Phase3 (±2 window, N≤5) results ({elapsed:.1f}s)")
print("=" * 70)
print(f"  N=2 matches: {stats['n2']} groups, {stats['n2_items']} items")
print(f"  N=3 matches: {stats['n3']} groups, {stats['n3_items']} items")
print(f"  N=4 matches: {stats['n4']} groups, {stats['n4_items']} items")
print(f"  N=5 matches: {stats['n5']} groups, {stats['n5_items']} items")
total_new = stats['n2_items'] + stats['n3_items'] + stats['n4_items'] + stats['n5_items']
print(f"  Total matched: {total_new} items")
print(f"  Remaining: B{len(remain_b)-len(all_matched_b)} + E{len(remain_e)-len(all_matched_e)} = {len(remain_b)+len(remain_e)-total_new} items")

# Overall stats
total_all = len(bank_txns) + len(ent_txns)
matched_all = p0 + p2 + p25 + total_new
print(f"\nOverall: {matched_all}/{total_all} = {matched_all/total_all*100:.1f}%")
