"""
银企对账测试脚本 - 验证日期显示 + Phase 2.5 匹配
模拟 bankReconciliation.ts 的核心逻辑，用 Python 端到端跑一次
"""
import json
import math
from collections import defaultdict
from datetime import date, datetime, timedelta
from openpyxl import load_workbook

# ===== 文件路径 =====
BANK_FILE = r"C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237流水.XLSX"
ENT_FILE  = r"C:\Users\CSCI\WPSDrive\6864036\WPS云盘\輪崗工作\银行对账\test data\new\01469910120237企业账.XLSX"

DATE_WINDOW_DAYS = 7

# ===== 模拟 parseDate (TS 版) =====
def excel_serial_to_date(serial):
    """模拟修复后的 excelSerialToDate: (serial - 25569) * 86400000"""
    return datetime(1970, 1, 1) + timedelta(days=serial - 25569)

def parse_date(val):
    """模拟 parseDate（纯数值路径，xlsx 用 openpyxl 读）"""
    if isinstance(val, datetime):
        return val
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime(val.year, val.month, val.day)
    if isinstance(val, (int, float)):
        if 30000 < val < 100000:
            return excel_serial_to_date(val)
    return None

def to_date_key(d):
    """模拟 toDateKey: 本地时间 YYYY-MM-DD"""
    return f"{d.year}-{d.month:02d}-{d.day:02d}"

def cents_equal(a, b):
    return round(a * 100) == round(b * 100)

def within_date_window(a, b, days=DATE_WINDOW_DAYS):
    return abs((a - b).total_seconds()) <= days * 86400

def parse_amount(s):
    """模拟 parseAmount"""
    if s is None:
        return 0.0
    s = str(s).replace(',', '').replace(' ', '').strip()
    if not s:
        return 0.0
    # 括号负数
    if s.startswith('(') and s.endswith(')'):
        s = '-' + s[1:-1]
    # 尾随负号
    if s.endswith('-'):
        s = '-' + s[:-1]
    # 尾随 DR/CR
    if s.upper().endswith('DR') or s.upper().endswith('CR'):
        s = s[:-2]
    try:
        return float(s)
    except ValueError:
        return 0.0

def clean_account(raw):
    s = str(raw).strip()
    if s.startswith("'") and len(s) > 1:
        s = s[1:]
    return s.strip()

# ===== 读 Excel =====
def read_xlsx(filepath, label):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        rows.append(list(row))
    
    print(f"\n{'='*60}")
    print(f"📄 {label}: {filepath}")
    print(f"   行数: {len(rows)} (含表头)")
    print(f"   表头: {rows[0]}")
    
    # 检查遇到的日期值
    date_samples = []
    for i, row in enumerate(rows[1:], start=2):
        for j, cell in enumerate(row):
            if isinstance(cell, (datetime, date)):
                date_samples.append((i, j, rows[0][j] if j < len(rows[0]) else f'Col{j}', cell, type(cell).__name__))
    if date_samples:
        print(f"   日期类型单元格 (前10个):")
        for s in date_samples[:10]:
            print(f"     行{s[0]} 列[{s[2]}]({s[1]}) = {s[3]} ({s[4]})")
    
    return rows, list(ws.iter_rows(min_row=1, values_only=False))

# ===== 手动逐 cell 读取（原始值，包括数字）=====
def read_xlsx_raw(filepath):
    """用 openpyxl 读原始值（number_format 等）"""
    wb = load_workbook(filepath, data_only=False)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1):
        cells = []
        for cell in row:
            cells.append({
                'value': cell.value,
                'type': type(cell.value).__name__,
                'number_format': cell.number_format
            })
        rows.append(cells)
    return rows

# ===== 主流程 =====
print("🔍 阶段 1: 原始数据检查")
bank_raw = read_xlsx_raw(BANK_FILE)
ent_raw = read_xlsx_raw(ENT_FILE)

# 检查日期列的值类型
def analyze_date_column(raw_rows, label):
    """分析日期列（假设第2列是日期）"""
    header = raw_rows[0]
    print(f"\n--- {label} 日期列分析 ---")
    
    # 找表头中包含 "日期" 的列
    date_cols = []
    for j, cell in enumerate(header):
        h = str(cell['value']).strip() if cell['value'] else ''
        if '日期' in h or 'date' in h.lower():
            date_cols.append(j)
    
    print(f"   日期相关列: {date_cols} (表头: {[header[j]['value'] for j in date_cols]})")
    
    for col_j in date_cols:
        print(f"\n   列 {col_j} ({header[col_j]['value']}):")
        types_count = defaultdict(int)
        number_formats = set()
        samples = []
        
        for i, row in enumerate(raw_rows[1:]):
            if col_j >= len(row):
                continue
            cell = row[col_j]
            v = cell['value']
            t = cell['type']
            nf = cell['number_format']
            types_count[t] += 1
            if nf:
                number_formats.add(nf)
            if len(samples) < 8 and v is not None:
                samples.append((i+2, v, t, nf))
        
        print(f"      类型分布: {dict(types_count)}")
        print(f"      Number formats: {number_formats}")
        print(f"      样本 (前8个):")
        for s in samples:
            print(f"        行{s[0]}: value={s[1]!r} type={s[2]} format={s[3]!r}")

analyze_date_column(bank_raw, "银行流水")
analyze_date_column(ent_raw, "企业账")

# ===== 阶段 2: data_only=True 读值，模拟 parseDate =====
print(f"\n{'='*60}")
print("🔍 阶段 2: data_only=True 读值 + parseDate 模拟")

def read_with_data_only(filepath):
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        rows.append(list(row))
    return rows

bank_data = read_with_data_only(BANK_FILE)
ent_data = read_with_data_only(ENT_FILE)

# 分析日期列
bank_header = bank_data[0]
ent_header = ent_data[0]

print(f"\n银行流水表头: {bank_header}")
print(f"企业账表头:   {ent_header}")

# 假设第2列(索引1)是日期
print("\n--- 银行流水日期解析测试 ---")
for i, row in enumerate(bank_data[1:30], start=2):
    val = row[1] if len(row) > 1 else None
    d = parse_date(val)
    key = to_date_key(d) if d else 'PARSE_FAIL'
    print(f"  行{i}: raw={val!r} ({type(val).__name__}) → {d} → key={key}")

print("\n--- 企业账日期解析测试 ---")
for i, row in enumerate(ent_data[1:30], start=2):
    val = row[1] if len(row) > 1 else None
    d = parse_date(val)
    key = to_date_key(d) if d else 'PARSE_FAIL'
    print(f"  行{i}: raw={val!r} ({type(val).__name__}) → {d} → key={key}")

# ===== 阶段 3: 模拟完整匹配流程 =====
print(f"\n{'='*60}")
print("🔍 阶段 3: 模拟完整银企对账匹配")

# 银行流水列配置 (需要根据实际表头调整)
# 典型 CSHK 银行流水: 账号 | 日期 | 摘要 | 收入 | 支出 | ...
# 企业账: 账号 | 日期 | 摘要 | 借方 | 贷方 | ...

# 先看看列数
print(f"银行流水的列数: {len(bank_header)}")
print(f"企业账的列数: {len(ent_header)}")

# 自动检测列索引
def find_col_idx(header, keywords):
    for i, h in enumerate(header):
        h_str = str(h).strip() if h else ''
        for kw in keywords:
            if kw in h_str:
                return i
    return -1

bank_acct_col = find_col_idx(bank_header, ['账号', '账户', '帐号'])
bank_date_col = find_col_idx(bank_header, ['日期'])
bank_income_col = find_col_idx(bank_header, ['收入', '贷方', '入账', 'CR'])
bank_expense_col = find_col_idx(bank_header, ['支出', '借方', '出账', 'DR'])

ent_acct_col = find_col_idx(ent_header, ['账号', '账户', '帐号', '科目'])
ent_date_col = find_col_idx(ent_header, ['日期'])
ent_debit_col = find_col_idx(ent_header, ['借方', '借', 'Debit'])
ent_credit_col = find_col_idx(ent_header, ['贷方', '贷', 'Credit'])

print(f"\n自动检测列索引:")
print(f"  银行: account={bank_acct_col}, date={bank_date_col}, income={bank_income_col}, expense={bank_expense_col}")
print(f"  企业: account={ent_acct_col}, date={ent_date_col}, debit={ent_debit_col}, credit={ent_credit_col}")

# 构建交易列表
class BankTx:
    def __init__(self, account, date, amount, direction):
        self.account = account
        self.date = date
        self.amount = amount
        self.direction = direction

class EntTx:
    def __init__(self, account, date, amount, direction):
        self.account = account
        self.date = date
        self.amount = amount
        self.direction = direction

bank_txns = []
for row in bank_data[1:]:
    acct = clean_account(row[bank_acct_col]) if bank_acct_col >= 0 and len(row) > bank_acct_col else ''
    if not acct:
        continue
    
    d = parse_date(row[bank_date_col]) if bank_date_col >= 0 and len(row) > bank_date_col else None
    if not d:
        continue
    
    income = parse_amount(row[bank_income_col]) if bank_income_col >= 0 and len(row) > bank_income_col else 0
    expense = parse_amount(row[bank_expense_col]) if bank_expense_col >= 0 and len(row) > bank_expense_col else 0
    
    if income != 0:
        bank_txns.append(BankTx(acct, d, income, '收入'))
    elif expense != 0:
        bank_txns.append(BankTx(acct, d, -abs(expense), '支出'))

ent_txns = []
for row in ent_data[1:]:
    acct = clean_account(row[ent_acct_col]) if ent_acct_col >= 0 and len(row) > ent_acct_col else ''
    if not acct:
        continue
    
    d = parse_date(row[ent_date_col]) if ent_date_col >= 0 and len(row) > ent_date_col else None
    if not d:
        continue
    
    debit = parse_amount(row[ent_debit_col]) if ent_debit_col >= 0 and len(row) > ent_debit_col else 0
    credit = parse_amount(row[ent_credit_col]) if ent_credit_col >= 0 and len(row) > ent_credit_col else 0
    
    if debit != 0:
        ent_txns.append(EntTx(acct, d, debit, '借方'))
    elif credit != 0:
        ent_txns.append(EntTx(acct, d, -credit, '貸方'))  # 标准SAP：贷方正数→取反

print(f"\n提取交易:")
print(f"  银行: {len(bank_txns)} 笔")
print(f"  企业: {len(ent_txns)} 笔")

# 按账号分组
bank_by_acct = defaultdict(list)
for t in bank_txns:
    bank_by_acct[t.account].append(t)
ent_by_acct = defaultdict(list)
for t in ent_txns:
    ent_by_acct[t.account].append(t)

bank_accts = sorted(set(t.account for t in bank_txns))
ent_accts = sorted(set(t.account for t in ent_txns))
overlap = [a for a in bank_accts if a in ent_accts]

print(f"\n账号分析:")
print(f"  银行账号: {bank_accts}")
print(f"  企业账号: {ent_accts}")
print(f"  交集: {overlap}")
print(f"  银行独有: {[a for a in bank_accts if a not in ent_accts]}")
print(f"  企业独有: {[a for a in ent_accts if a not in bank_accts]}")

# ===== 阶段 4: 逐账号模拟匹配 =====
print(f"\n{'='*60}")
print("🔍 阶段 4: 逐账号模拟 Phase 0→2.5 匹配")

for acct in overlap:
    banks = sorted(bank_by_acct[acct], key=lambda t: t.date)
    ents = sorted(ent_by_acct[acct], key=lambda t: t.date)
    
    print(f"\n--- 账号: {acct} ---")
    print(f"  银行流水: {len(banks)} 笔")
    print(f"  企业账:   {len(ents)} 笔")
    
    # 按方向统计
    b_income = [t for t in banks if t.direction == '收入']
    b_expense = [t for t in banks if t.direction == '支出']
    e_debit = [t for t in ents if t.direction == '借方']
    e_credit = [t for t in ents if t.direction == '貸方']
    
    b_inc_sum = sum(t.amount for t in b_income)
    b_exp_sum = sum(t.amount for t in b_expense)
    e_deb_sum = sum(t.amount for t in e_debit)
    e_cre_sum = sum(t.amount for t in e_credit)
    
    print(f"  收入: {len(b_income)}笔 sum={b_inc_sum:.2f} | 借方: {len(e_debit)}笔 sum={e_deb_sum:.2f}")
    print(f"  支出: {len(b_expense)}笔 sum={b_exp_sum:.2f} | 贷方: {len(e_credit)}笔 sum={e_cre_sum:.2f}")
    
    # Phase 0 快速通道
    fast_income = cents_equal(b_inc_sum, e_deb_sum) and len(b_income) > 0 and len(e_debit) > 0
    fast_expense = cents_equal(b_exp_sum, e_cre_sum) and len(b_expense) > 0 and len(e_credit) > 0
    if fast_income:
        print(f"  ⚡ Phase 0 快速通道(收入↔借方): 触发! sum差={round(b_inc_sum*100)-round(e_deb_sum*100)}分")
    if fast_expense:
        print(f"  ⚡ Phase 0 快速通道(支出↔贷方): 触发! sum差={round(b_exp_sum*100)-round(e_cre_sum*100)}分")
    
    if not fast_income and len(b_income) > 0 and len(e_debit) > 0:
        diff = round(b_inc_sum*100) - round(e_deb_sum*100)
        print(f"  ❌ 快速通道(收入↔借方): 未触发, 差额={diff}分 (={diff/100:.2f}元)")
    if not fast_expense and len(b_expense) > 0 and len(e_credit) > 0:
        diff = round(b_exp_sum*100) - round(e_cre_sum*100)
        print(f"  ❌ 快速通道(支出↔贷方): 未触发, 差额={diff}分 (={diff/100:.2f}元)")
    
    # Phase 1: 模拟 1:1 按金额匹配
    # 简化版：用 hashtable
    ent_by_cents = defaultdict(list)
    for i, e in enumerate(ents):
        c = round(e.amount * 100)
        ent_by_cents[c].append((i, e))
    
    matched_1to1 = 0
    used_ent = set()
    used_bank = set()
    
    for bi, b in enumerate(banks):
        if bi in used_bank:
            continue
        c = round(b.amount * 100)
        bucket = ent_by_cents.get(c, [])
        for ei, e in bucket:
            if ei not in used_ent:
                if within_date_window(b.date, e.date):
                    used_bank.add(bi)
                    used_ent.add(ei)
                    matched_1to1 += 1
                    break
        else:
            # 兜底：金额一致但日期窗口外
            for ei, e in bucket:
                if ei not in used_ent:
                    used_bank.add(bi)
                    used_ent.add(ei)
                    matched_1to1 += 1
                    break
    
    print(f"  Phase 2 (1:1): {matched_1to1} 对匹配")
    
    # Phase 2.5: 日期分桶快速通道
    remaining_bank = [t for i, t in enumerate(banks) if i not in used_bank]
    remaining_ent = [t for i, t in enumerate(ents) if i not in used_ent]
    
    bank_by_date = defaultdict(list)
    for t in remaining_bank:
        bank_by_date[to_date_key(t.date)].append(t)
    ent_by_date = defaultdict(list)
    for t in remaining_ent:
        ent_by_date[to_date_key(t.date)].append(t)
    
    date_bucket_matched = 0
    for dk, b_items in bank_by_date.items():
        e_items = ent_by_date.get(dk, [])
        if not e_items:
            continue
        
        # 收入方向
        b_inc = [t for t in b_items if t.amount > 0]
        e_deb = [t for t in e_items if t.amount > 0 and t.direction == '借方']
        if b_inc and e_deb:
            bs = sum(t.amount for t in b_inc)
            es = sum(t.amount for t in e_deb)
            if cents_equal(bs, es):
                date_bucket_matched += len(b_inc) + len(e_deb)
                # 标记为已用（简化）
        
        # 支出方向
        b_exp = [t for t in b_items if t.amount < 0]
        e_cre = [t for t in e_items if t.amount < 0 and t.direction == '貸方']
        if b_exp and e_cre:
            bs = sum(t.amount for t in b_exp)
            es = sum(t.amount for t in e_cre)
            if cents_equal(bs, es):
                date_bucket_matched += len(b_exp) + len(e_cre)
    
    print(f"  Phase 2.5 (日期分桶): {date_bucket_matched} 条匹配")
    print(f"  剩余: 银行{len(remaining_bank)}笔, 企业{len(remaining_ent)}笔")
    
    # 如果剩余不多，列出详情
    if len(remaining_bank) <= 10 and len(remaining_ent) <= 10:
        print(f"\n  剩余银行流水:")
        for t in remaining_bank:
            print(f"    {to_date_key(t.date)} {t.direction} {t.amount:>12.2f}")
        print(f"  剩余企业账:")
        for t in remaining_ent:
            print(f"    {to_date_key(t.date)} {t.direction} {t.amount:>12.2f}")

print(f"\n{'='*60}")
print("✅ 测试完成")
